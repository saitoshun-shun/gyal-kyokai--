"""
事業計画書 - バズ社PL Year1月次（❶大成功パターン）v2
収益3本柱:
  ① TikTokメディアタイアップ（インバウンドプロモ企業向け）
  ② IGメディアタイアップ（インバウンドプロモ企業向け）
  ③ TikTok Shop / WESELL EC（ギャルグッズ販売）
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Year1月次PL_大成功"

# ─── カラー（シンプル） ─────────────────────
NAVY    = "1A1A2E"
NAVY2   = "16213E"
BLUE    = "2E4A7A"
INDIGO  = "3B4D8E"
RED_ACC = "C0392B"
GREEN   = "1E8449"
LT_BLUE = "D6EAF8"
LT_GRN  = "D5F5E3"
LT_GRN2 = "EAFAF1"
LT_ORG  = "FEF9E7"
LT_GRY  = "F2F3F4"
MID_GRY = "E5E7E9"
WHITE   = "FFFFFF"
GRY_TXT = "7F8C8D"
PURPLE  = "6C3483"
LT_PUR  = "F4ECF7"

# ─── ヘルパー ──────────────────────────────
def c(row, col, val="", bold=False, sz=10, fg="222222", bg=None,
      ha="center", wrap=False, italic=False, fmt=None, indent=0):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(name="Meiryo UI", bold=bold, size=sz,
                     color=fg, italic=italic)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=ha, vertical="center",
                               wrap_text=wrap, indent=indent)
    if fmt:
        cell.number_format = fmt
    return cell

def fill_empty(row, col_start, col_end, bg):
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=bg)

# ─── レイアウト ──────────────────────────
NCOL_ITEM  = 1
NCOL_M1    = 2
NCOL_TOTAL = 14
MONTHS = [f"M{i}" for i in range(1, 13)]

ws.column_dimensions["A"].width = 36
for col in range(2, 15):
    ws.column_dimensions[get_column_letter(col)].width = 10

# ─────────────────────────────────────────────────
# 数値定義（大成功パターン）
# ─────────────────────────────────────────────────

# ① TikTokメディアタイアップ（万円）
#   企業がTKアカウントに掲載するタイアップフィー。M3から受注開始。
tk_tie = [0, 0, 100, 150, 200, 300, 350, 400, 500, 500, 600, 700]

# ② IGメディアタイアップ（万円）
#   企業がIGアカウントに掲載するタイアップフィー。M4から受注開始。
ig_tie = [0, 0, 0, 100, 150, 200, 300, 350, 400, 500, 500, 600]

tie_total = [t + i for t, i in zip(tk_tie, ig_tie)]  # タイアップ合計

# ③ TikTok Shop / WESELL EC グロス売上（万円）
ec_units = [0, 0, 50, 120, 250, 450, 700, 1000, 1300, 1700, 2200, 2800]
ec_atp   = [0, 0, 5000, 5500, 6000, 6500, 7000, 7200, 7500, 7800, 8000, 8200]
ec_gross = [round(u * a / 10000) for u, a in zip(ec_units, ec_atp)]

rev_total = [t + e for t, e in zip(tie_total, ec_gross)]

# ─── 原価 ────────────────────────────────
# タイアップ原価（シンプル構造）
#   ギャル協会ライセンス料: タイアップ売上 × 20%（出演・監修・IPライセンス料）
#   コンテンツ制作費（外注）: タイアップ売上 × 10%（撮影・編集等）
#   ⇒ 原価率30%、粗利率70%
tie_license = [round(t * 0.20) for t in tie_total]   # ライセンス料
tie_prod    = [round(t * 0.10) for t in tie_total]   # 制作費
tie_cogs    = [l + p for l, p in zip(tie_license, tie_prod)]
tie_gross   = [t - cg for t, cg in zip(tie_total, tie_cogs)]

# EC原価（商品販売のコスト構造）
#   商品仕入れ原価: EC売上 × 45%
#   TikTok Shop手数料: EC売上 × 6%
#   ギャル協会RS（物販）: EC売上 × 15%
#   ⇒ 原価率66%、粗利率34%
ec_cogs_item = [round(e * 0.45) for e in ec_gross]   # 商品仕入れ
ec_fee       = [round(e * 0.06) for e in ec_gross]   # TikTok Shop手数料
ec_gal_rs    = [round(e * 0.15) for e in ec_gross]   # ギャル協会RS（物販）
ec_cogs      = [ci + f + g for ci, f, g in zip(ec_cogs_item, ec_fee, ec_gal_rs)]
ec_gross_pft = [e - ec for e, ec in zip(ec_gross, ec_cogs)]  # EC粗利

# 合算
cogs_total  = [t + e for t, e in zip(tie_cogs, ec_cogs)]
gross_total = [r - cg for r, cg in zip(rev_total, cogs_total)]

# ─── 販管費 ──────────────────────────────
ad_exp = [100,150,150,150,200,200,250,250,250,250,250,250]  # 広告費
labor  = [100,100,100,100,100,100,100,100,100,100,100,100]  # 人件費（バズ社員）
other  = [50, 50, 50, 60, 60, 60, 70, 70, 80, 80, 90, 90]  # その他
sga    = [a + l + o for a, l, o in zip(ad_exp, labor, other)]

# ─── 営業利益 ────────────────────────────
op_profit = [g - s for g, s in zip(gross_total, sga)]
bep_month = next((i for i, v in enumerate(op_profit) if v > 0), None)

# 参考比率
gp_rate = [round(g / r * 100, 1) if r > 0 else 0.0
           for g, r in zip(gross_total, rev_total)]
tie_gp_rate = [round(tg / t * 100, 1) if t > 0 else 0.0
               for tg, t in zip(tie_gross, tie_total)]
ec_gp_rate  = [round(eg / e * 100, 1) if e > 0 else 0.0
               for eg, e in zip(ec_gross_pft, ec_gross)]
op_rate = [round(o / r * 100, 1) if r > 0 else 0.0
           for o, r in zip(op_profit, rev_total)]

# ─────────────────────────────────────────────────
# タイアップ件数KPI
# ─────────────────────────────────────────────────
tk_cases = [0, 0, 1, 1, 2, 2, 3, 3, 4, 4, 5, 5]
ig_cases = [0, 0, 0, 1, 1, 2, 2, 3, 3, 4, 4, 5]
tk_atp   = [0, 0, 100, 150, 100, 150, 117, 133, 125, 125, 120, 140]  # 平均単価
ig_atp   = [0, 0, 0,   100, 150, 100, 150, 117, 133, 125, 125, 120]

# ─────────────────────────────────────────────────
# タイトル
# ─────────────────────────────────────────────────
ROW = 1
ws.row_dimensions[ROW].height = 38
ws.merge_cells(f"A{ROW}:N{ROW}")
cell = ws.cell(row=ROW, column=1,
               value="❶ 大成功パターン  ／  Year1 月次事業計画（バズ社PL）")
cell.font = Font(name="Meiryo UI", bold=True, size=16, color=WHITE)
cell.fill = PatternFill("solid", fgColor=NAVY)
cell.alignment = Alignment(horizontal="center", vertical="center")

ROW += 1
ws.row_dimensions[ROW].height = 16
ws.merge_cells(f"A{ROW}:N{ROW}")
cell = ws.cell(row=ROW, column=1,
               value="ギャル×インバウンド共同事業（TK・IGタイアップメディア × TikTok Shop WESELL）　2026年4月〜2027年3月　単位：万円")
cell.font = Font(name="Meiryo UI", size=9, color="CCCCCC")
cell.fill = PatternFill("solid", fgColor=NAVY2)
cell.alignment = Alignment(horizontal="center", vertical="center")

# ─────────────────────────────────────────────────
# ヘッダー行
# ─────────────────────────────────────────────────
ROW += 1
ws.row_dimensions[ROW].height = 26
c(ROW, NCOL_ITEM, "項目", bold=True, sz=10, fg=WHITE, bg=BLUE)
for i, m in enumerate(MONTHS):
    c(ROW, NCOL_M1+i, m, bold=True, sz=10, fg=WHITE, bg=BLUE)
c(ROW, NCOL_TOTAL, "Year1合計", bold=True, sz=10, fg=WHITE, bg=BLUE)
HDR_ROW = ROW

# ─────────────────────────────────────────────────
# KPIセクション
# ─────────────────────────────────────────────────
ROW += 1
ws.row_dimensions[ROW].height = 22
ws.merge_cells(f"A{ROW}:N{ROW}")
cell = ws.cell(row=ROW, column=1, value="▌ 重要KPI（月次推移）")
cell.font = Font(name="Meiryo UI", bold=True, size=11, color=WHITE)
cell.fill = PatternFill("solid", fgColor=BLUE)
cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

kpi_rows = [
    # (ラベル, データ, 書式, ★KPI, 集計方法)
    ("TikTokフォロワー数（累計）",
     [500,2000,5000,10000,18000,28000,40000,55000,72000,92000,115000,140000],
     "#,##0", True, "last"),
    ("IGフォロワー数（累計）",
     [300,1200,3000,6000,12000,20000,30000,43000,58000,75000,95000,120000],
     "#,##0", True, "last"),
    ("TKタイアップ受注件数（月次）",
     tk_cases, "#,##0", True, "sum"),
    ("TKタイアップ平均単価（万円）",
     tk_atp, "#,##0", False, "avg_nz"),
    ("IGタイアップ受注件数（月次）",
     ig_cases, "#,##0", True, "sum"),
    ("IGタイアップ平均単価（万円）",
     ig_atp, "#,##0", False, "avg_nz"),
    ("EC注文件数（TikTok Shop）",
     ec_units, "#,##0", True, "sum"),
    ("EC平均注文単価（円）",
     ec_atp, "#,##0", False, "avg_nz"),
]

for ki, (label, vals, fmt, is_kpi, agg) in enumerate(kpi_rows):
    ROW += 1
    ws.row_dimensions[ROW].height = 17
    row_bg = LT_GRY if ki % 2 == 0 else WHITE

    cell = ws.cell(row=ROW, column=NCOL_ITEM, value=label)
    cell.font = Font(name="Meiryo UI", size=9, bold=is_kpi,
                     color=BLUE if is_kpi else "333333")
    cell.fill = PatternFill("solid", fgColor=row_bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    for j, v in enumerate(vals):
        cell2 = ws.cell(row=ROW, column=NCOL_M1+j, value=v if v != 0 else "-")
        cell2.font = Font(name="Meiryo UI", size=9, bold=is_kpi,
                          color=BLUE if is_kpi else "444444")
        cell2.fill = PatternFill("solid", fgColor=row_bg)
        cell2.alignment = Alignment(horizontal="center", vertical="center")
        if v != 0:
            cell2.number_format = fmt

    # 集計
    non_z = [v for v in vals if v != 0]
    if agg == "last":    agg_v = vals[-1]
    elif agg == "sum":   agg_v = sum(vals)
    elif agg == "avg_nz":agg_v = round(sum(non_z)/len(non_z)) if non_z else 0
    else:                agg_v = sum(vals)

    cell3 = ws.cell(row=ROW, column=NCOL_TOTAL, value=agg_v if agg_v != 0 else "-")
    cell3.font = Font(name="Meiryo UI", size=9, bold=True,
                      color=BLUE if is_kpi else "333333")
    cell3.fill = PatternFill("solid", fgColor=row_bg)
    cell3.alignment = Alignment(horizontal="center", vertical="center")
    if agg_v != 0:
        cell3.number_format = fmt

# ─────────────────────────────────────────────────
# PLセクション
# ─────────────────────────────────────────────────
ROW += 1
ws.row_dimensions[ROW].height = 22
ws.merge_cells(f"A{ROW}:N{ROW}")
cell = ws.cell(row=ROW, column=1, value="▌ 損益計算書（P/L）　単位：万円")
cell.font = Font(name="Meiryo UI", bold=True, size=11, color=WHITE)
cell.fill = PatternFill("solid", fgColor=NAVY)
cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# ─── PL行定義 ───────────────────────────────────
#  style: "sec" / "normal" / "sub" / "gross" / "rate" / "op" / "blank" / "ec_gross" / "ec_sub"

pl_rows = [
    # ── 売上 ──
    # (ラベル, データ, style, 合計方法, 書式)
    # 合計方法: "sum" / "avg_nz" / None
    ("【売上高】",                             None,     "sec",       None,      "#,##0"),
    ("  ① TikTokメディアタイアップ",          tk_tie,   "normal",    "sum",     "#,##0"),
    ("     受注社数（社）",                    tk_cases, "kpi_sub",   "sum",     "#,##0"),
    ("     平均単価（万円/社）",               tk_atp,   "kpi_sub",   "avg_nz",  "#,##0"),
    ("  ② IGメディアタイアップ",              ig_tie,   "normal",    "sum",     "#,##0"),
    ("     受注社数（社）",                    ig_cases, "kpi_sub",   "sum",     "#,##0"),
    ("     平均単価（万円/社）",               ig_atp,   "kpi_sub",   "avg_nz",  "#,##0"),
    ("  タイアップ売上合計",                   tie_total,"sub",       "sum",     "#,##0"),
    ("  ③ TikTok Shop / WESELL EC売上",       ec_gross, "normal",    "sum",     "#,##0"),
    ("     注文件数（件）",                    ec_units, "kpi_sub",   "sum",     "#,##0"),
    ("     平均注文単価（円）",                ec_atp,   "kpi_sub",   "avg_nz",  "#,##0"),
    ("売上合計",                               rev_total,"total_rev", "sum",     "#,##0"),
    ("", None, "blank", None, None),

    # ── タイアップ原価 ──
    ("【タイアップ原価】※ほぼライセンス料のみ", None,        "sec",       None,     "#,##0"),
    ("  ギャル協会ライセンス料（タイアップ×20%）", tie_license, "normal",  "sum",    "#,##0"),
    ("  コンテンツ制作費（タイアップ×10%）",    tie_prod,    "normal",    "sum",    "#,##0"),
    ("  タイアップ原価合計",                    tie_cogs,    "sub",       "sum",    "#,##0"),
    ("  タイアップ粗利",                        tie_gross,   "tie_gross", "sum",    "#,##0"),
    ("  タイアップ粗利率",                      tie_gp_rate, "rate",      "avg_nz", None),
    ("", None, "blank", None, None),

    # ── EC原価 ──
    ("【TikTok Shop EC原価】",                  None,         "sec",       None,    "#,##0"),
    ("  商品仕入れ原価（EC売上×45%）",          ec_cogs_item, "normal",    "sum",   "#,##0"),
    ("  TikTok Shop手数料（EC売上×6%）",        ec_fee,       "normal",    "sum",   "#,##0"),
    ("  ギャル協会RS・物販（EC売上×15%）",      ec_gal_rs,    "normal",    "sum",   "#,##0"),
    ("  EC原価合計",                            ec_cogs,      "sub",       "sum",   "#,##0"),
    ("  TikTok Shop 粗利",                      ec_gross_pft, "ec_gross",  "sum",   "#,##0"),
    ("  TikTok Shop 粗利率",                    ec_gp_rate,   "rate",      "avg_nz",None),
    ("", None, "blank", None, None),

    # ── 合算 ──
    ("売上原価合計",                            cogs_total,  "total_cost","sum",    "#,##0"),
    ("粗利益合計",                              gross_total, "gross",     "sum",    "#,##0"),
    ("粗利率（合計）",                          gp_rate,     "rate",      "avg_nz", None),
    ("", None, "blank", None, None),

    # ── 販管費 ──
    ("【販売管理費】",                          None,    "sec",       None,  "#,##0"),
    ("  広告費（TK/IG広告運用）",               ad_exp,  "normal",    "sum", "#,##0"),
    ("  人件費（バズ社員 ／ ENTIALはゼロ）",    labor,   "normal",    "sum", "#,##0"),
    ("  その他販管費",                          other,   "normal",    "sum", "#,##0"),
    ("販売管理費合計",                          sga,     "total_cost","sum", "#,##0"),
    ("", None, "blank", None, None),

    # ── 利益 ──
    ("営業利益",                                op_profit, "op",   "sum",    "#,##0"),
    ("営業利益率",                              op_rate,   "rate", "avg_nz", None),
]

STYLE = {
    "sec":       dict(bold=True,  sz=9,  fg="1A1A2E", bg=MID_GRY,  ha="left"),
    "normal":    dict(bold=False, sz=9,  fg="333333", bg=WHITE,     ha="left"),
    "kpi_sub":   dict(bold=False, sz=8,  fg="888888", bg="FAFAFA",  ha="left"),
    "sub":       dict(bold=True,  sz=9,  fg="1A3A6B", bg=LT_BLUE,   ha="left"),
    "total_rev": dict(bold=True,  sz=10, fg="1A3A6B", bg=LT_BLUE,   ha="left"),
    "total_cost":dict(bold=True,  sz=9,  fg="555555", bg=LT_GRY,    ha="left"),
    "tie_gross": dict(bold=True,  sz=10, fg="4A235A", bg=LT_PUR,    ha="left"),
    "ec_gross":  dict(bold=True,  sz=10, fg="145A32", bg=LT_GRN2,   ha="left"),
    "gross":     dict(bold=True,  sz=10, fg="145A32", bg=LT_GRN,    ha="left"),
    "rate":      dict(bold=False, sz=9,  fg=GRY_TXT,  bg=LT_GRY,   ha="left"),
    "op":        dict(bold=True,  sz=11, fg=WHITE,     bg=NAVY,     ha="left"),
    "blank":     dict(bold=False, sz=6,  fg=WHITE,     bg=WHITE,    ha="left"),
}

for pi, (label, vals, style, agg_method, num_fmt) in enumerate(pl_rows):
    ROW += 1
    ws.row_dimensions[ROW].height = (6 if style == "blank"
                                     else 15 if style == "kpi_sub"
                                     else 18)
    st = STYLE[style]

    cell = ws.cell(row=ROW, column=NCOL_ITEM, value=label)
    cell.font = Font(name="Meiryo UI", bold=st["bold"], size=st["sz"], color=st["fg"])
    cell.fill = PatternFill("solid", fgColor=st["bg"])
    cell.alignment = Alignment(horizontal=st["ha"], vertical="center", indent=1)

    if vals is None:
        fill_empty(ROW, NCOL_M1, NCOL_TOTAL, st["bg"])
        continue

    # データ書式の決定
    is_rate = (style == "rate")
    is_kpi  = (style == "kpi_sub")
    cell_fmt = ('0.0"%"' if is_rate else (num_fmt or "#,##0"))

    for j, v in enumerate(vals):
        col = NCOL_M1 + j
        # ゼロは "-" 表示（kpi_subと rate以外でもゼロは "-"）
        disp_val = "-" if (v == 0 and style in ("kpi_sub",)) else v
        cell2 = ws.cell(row=ROW, column=col, value=disp_val)

        # BEP月の強調（営業利益行のみ）
        if style == "op" and j == bep_month:
            cell2.fill = PatternFill("solid", fgColor=GREEN)
            cell2.font = Font(name="Meiryo UI", bold=True, size=st["sz"], color=WHITE)
        else:
            cell2.fill = PatternFill("solid", fgColor=st["bg"])
            fg_v = (WHITE if style == "op" else
                    ("145A32" if (style in ("gross","ec_gross","tie_gross") and v >= 0)
                     else RED_ACC if (style in ("gross","ec_gross","tie_gross") and v < 0)
                     else st["fg"]))
            cell2.font = Font(name="Meiryo UI", bold=st["bold"], size=st["sz"], color=fg_v)

        cell2.alignment = Alignment(horizontal="center", vertical="center")
        if disp_val != "-":
            cell2.number_format = cell_fmt

    # ── 合計列 ──
    non_z2 = [v for v in vals if v != 0]
    if agg_method == "avg_nz":
        agg_v  = round(sum(non_z2)/len(non_z2), 1) if non_z2 else 0.0
        fmt2   = '0.0"%"' if is_rate else (num_fmt or "#,##0")
    else:
        agg_v = sum(vals)
        fmt2  = '0.0"%"' if is_rate else (num_fmt or "#,##0")

    disp_agg = "-" if (agg_v == 0 and style == "kpi_sub") else agg_v
    cell3 = ws.cell(row=ROW, column=NCOL_TOTAL, value=disp_agg)
    fg3 = (WHITE if style == "op" else
           ("145A32" if (style in ("gross","ec_gross","tie_gross") and agg_v >= 0)
            else RED_ACC if (style in ("gross","ec_gross","tie_gross") and agg_v < 0)
            else st["fg"]))
    bg3 = (GREEN if (style == "op" and agg_v >= 0)
           else RED_ACC if (style == "op" and agg_v < 0)
           else st["bg"])
    cell3.font = Font(name="Meiryo UI", bold=True, size=st["sz"], color=fg3)
    cell3.fill = PatternFill("solid", fgColor=bg3)
    cell3.alignment = Alignment(horizontal="center", vertical="center")
    if disp_agg != "-":
        cell3.number_format = fmt2

# ─────────────────────────────────────────────────
# BEP注記
# ─────────────────────────────────────────────────
ROW += 1
ws.row_dimensions[ROW].height = 18
ws.merge_cells(f"A{ROW}:N{ROW}")
bep_label = f"M{bep_month+1}" if bep_month is not None else "Year1内なし"
cell = ws.cell(row=ROW, column=1,
               value=(f"★ 損益分岐：{bep_label} から黒字転換（営業利益行の緑セル）　"
                      f"Year1通期営業利益：{sum(op_profit):,}万円　"
                      f"Year1売上：{sum(rev_total):,}万円"))
cell.font = Font(name="Meiryo UI", size=9, bold=True, color=NAVY)
cell.fill = PatternFill("solid", fgColor=LT_GRN)
cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# ─────────────────────────────────────────────────
# 注記フッター
# ─────────────────────────────────────────────────
ROW += 1
ws.row_dimensions[ROW].height = 14
ws.merge_cells(f"A{ROW}:N{ROW}")
cell = ws.cell(row=ROW, column=1,
               value=("【注記】 単位：万円　"
                      "／ タイアップ原価：ライセンス料20%＋制作費10%＝計30%　"
                      "／ EC原価：商品仕入45%＋TK手数料6%＋ギャル協会RS15%＝計66%　"
                      "／ ENTIALへの費用はバズPLに計上なし（別途協議）"))
cell.font = Font(name="Meiryo UI", size=8, color=GRY_TXT, italic=True)
cell.fill = PatternFill("solid", fgColor=LT_GRY)
cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# ─────────────────────────────────────────────────
# 全セルにborder
# ─────────────────────────────────────────────────
thin = Side(style="thin", color="DDDDDD")
for row in ws.iter_rows(min_row=HDR_ROW, max_row=ROW,
                        min_col=1, max_col=NCOL_TOTAL):
    for cell in row:
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

ws.freeze_panes = "B4"

OUTPUT = "PL_Year1月次_大成功_v2.xlsx"
wb.save(OUTPUT)

print(f"✅ 完成: {OUTPUT}")
print()
print("── Year1通期サマリー ──")
print(f"  売上合計          : {sum(rev_total):>8,} 万円")
print(f"    うちタイアップ  : {sum(tie_total):>8,} 万円")
print(f"    うちEC          : {sum(ec_gross):>8,} 万円")
print()
print(f"  タイアップ粗利    : {sum(tie_gross):>8,} 万円"
      f"  （粗利率 {round(sum(tie_gross)/sum(tie_total)*100,1)}%）")
print(f"  TikTok Shop粗利  : {sum(ec_gross_pft):>8,} 万円"
      f"  （粗利率 {round(sum(ec_gross_pft)/sum(ec_gross)*100,1)}%）")
print(f"  粗利合計          : {sum(gross_total):>8,} 万円"
      f"  （粗利率 {round(sum(gross_total)/sum(rev_total)*100,1)}%）")
print()
print(f"  販管費合計        : {sum(sga):>8,} 万円")
print(f"  営業利益          : {sum(op_profit):>8,} 万円")
print(f"  損益分岐点        : M{bep_month+1}" if bep_month is not None
      else "  損益分岐点        : Year1内なし")
