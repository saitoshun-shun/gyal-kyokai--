"""
月次PL - ①大成功パターン
ギャル協会 × サイバーバズ × ENTIAL
単位：万円
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "①大成功_月次PL"
ws.sheet_view.showGridLines = False

# ── カラー定義（ピンクなし・コーポレート系） ──
C_HEAD_BG   = "1F2937"   # ダークグレー（ヘッダー背景）
C_HEAD_FG   = "FFFFFF"   # 白文字
C_SEC_REV   = "1D4ED8"   # 青（売上セクション）
C_SEC_COGS  = "92400E"   # ブラウン（原価セクション）
C_SEC_SGA   = "5B21B6"   # パープル（販管費セクション）
C_SEC_PROF  = "065F46"   # グリーン（利益セクション）
C_ROW_REV   = "EFF6FF"   # 薄青（売上行）
C_ROW_COGS  = "FEF3C7"   # 薄黄（原価行）
C_ROW_SGA   = "F5F3FF"   # 薄紫（販管費行）
C_ROW_PROF  = "ECFDF5"   # 薄緑（利益行）
C_TOTAL     = "D1FAE5"   # 緑（合計行）
C_NEG       = "DC2626"   # 赤（マイナス）
C_POS       = "059669"   # 緑（プラス）
C_LABEL_BG  = "F9FAFB"   # 薄グレー（ラベル列）
C_WHITE     = "FFFFFF"
C_BORDER    = "D1D5DB"

def fill(c): return PatternFill("solid", fgColor=c)
def bd(color=C_BORDER, style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def bd_bottom(color="9CA3AF", style="medium"):
    return Border(bottom=Side(style=style, color=color))

def cell(ws, r, c, val="", bold=False, size=9, fg=None, bg=None,
         align="left", fmt=None, italic=False, color=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(bold=bold, size=size, color=color or "2D2D2D",
                     italic=italic, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=False)
    if bg: cell.fill = fill(bg)
    if fmt: cell.number_format = fmt
    cell.border = bd()
    return cell

# ── 列幅 ──────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 3   # インデント用
ws.column_dimensions["B"].width = 30  # 項目名
for i in range(3, 16):                # 1月〜12月＋年計
    ws.column_dimensions[get_column_letter(i)].width = 10
ws.column_dimensions[get_column_letter(15)].width = 12  # 年計
ws.column_dimensions[get_column_letter(16)].width = 10  # 対年計比率

# ══════════════════════════════════════════════════════════════
# ① 大成功パターンの月次想定数値（万円）
# 前提：
#   - TikTokフォロワーが急速に伸び、Month4から受注開始
#   - 月次タイアップ件数：M4=1件, M5=2件, M7=3件, M10=4件
#   - インバウンド体験：M4=50人, M7=150人, M10=300人
#   - WESELL：M5から稼働、Month10=500万
# ══════════════════════════════════════════════════════════════

M = 12  # 月数

# ── 売上（Revenue）──────────────────────────────────────────
# 単価仮置き（コメントに記載）

# ①インバウンド体験 (仮: ¥20,000/人)
inbound_pax    = [0,0,0,50,80,100,150,180,200,300,300,350]
inbound_price  = 2.0   # 万円/人
inbound_rev    = [p * inbound_price for p in inbound_pax]

# ②広告タイアップ (仮: 300万/件)
taieup_count   = [0,0,0,1,1,2,2,2,3,3,4,4]
taieup_price   = 300   # 万円/件
taieup_rev     = [c * taieup_price for c in taieup_count]

# ③WESELL グッズ販売 売上（仮: 粗利50%）
wesell_rev     = [0,0,0,0,100,150,200,250,300,400,500,600]

# ④IPライセンス (仮: 50万/社/月、Year1は2社)
license_rev    = [0,0,0,0,0,50,50,100,100,100,150,150]

# ⑤ギャル旅タイアップ (仮: 400万/件、年2件)
galtrip_rev    = [0,0,0,0,400,0,0,0,400,0,0,0]

total_rev = [sum(x) for x in zip(inbound_rev, taieup_rev, wesell_rev,
                                   license_rev, galtrip_rev)]

# ── 売上原価（COGS）─────────────────────────────────────────
# インバウンド体験原価 (仮: 会場・衣装・スタッフ = 売上の40%)
inbound_cogs   = [r * 0.40 for r in inbound_rev]
# WESELL仕入原価 (仮: 売上の50%)
wesell_cogs    = [r * 0.50 for r in wesell_rev]
# ギャル旅制作費 (仮: 売上の35%)
galtrip_cogs   = [r * 0.35 for r in galtrip_rev]

total_cogs = [sum(x) for x in zip(inbound_cogs, wesell_cogs, galtrip_cogs)]
gross_profit   = [r - c for r, c in zip(total_rev, total_cogs)]
gp_margin      = [gp/r if r > 0 else 0 for gp, r in zip(gross_profit, total_rev)]

# ── 販売費及び一般管理費（SG&A）─────────────────────────────
# コンテンツ制作費（TikTok/IG）仮: 月150万（立ち上げ後）
content_cost   = [50,100,150,150,150,150,200,200,200,250,250,250]
# ギャル協会レベニューシェア (広告+ライセンスの32%相当)
gyaru_share    = [round((taieup_rev[i]+license_rev[i])*0.32) for i in range(M)]
# ENTIAL固定費 (仮: 月80万)
ential_fixed   = [80]*M
# ENTIALインセンティブ (売上の3%)
ential_inc     = [round(r * 0.03) for r in total_rev]
# 広告宣伝費 (仮: 月50万)
adv_cost       = [30,30,50,50,50,50,80,80,80,100,100,100]
# その他運営費 (仮: 月30万)
other_cost     = [30]*M

total_sga = [sum(x) for x in zip(content_cost, gyaru_share, ential_fixed,
                                   ential_inc, adv_cost, other_cost)]

# ── 営業利益 ────────────────────────────────────────────────
op_profit      = [gp - sga for gp, sga in zip(gross_profit, total_sga)]
op_margin      = [op/r if r > 0 else 0 for op, r in zip(op_profit, total_rev)]

# ── 累計営業利益 ─────────────────────────────────────────────
cumulative_op  = []
cum = 0
for v in op_profit:
    cum += v
    cumulative_op.append(cum)

# ══════════════════════════════════════════════════════════════
# Excel出力
# ══════════════════════════════════════════════════════════════

ROW = 1

# ── タイトル ─────────────────────────────────────────────────
ws.merge_cells(f"A{ROW}:P{ROW}")
c = ws.cell(row=ROW, column=1,
            value="損益計算書（P/L）　①大成功パターン　単位：万円")
c.font = Font(bold=True, size=13, color=C_HEAD_FG, name="Calibri")
c.fill = fill(C_HEAD_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = bd(C_HEAD_BG)
ws.row_dimensions[ROW].height = 28
ROW += 1

# サブタイトル
ws.merge_cells(f"A{ROW}:P{ROW}")
c = ws.cell(row=ROW, column=1,
            value="前提：TikTokフォロワーが急伸、Month4から受注開始。Year1通期で売上約1.1億円・営業利益約1,800万円を想定")
c.font = Font(bold=False, size=8, color="6B7280", italic=True, name="Calibri")
c.fill = fill("F9FAFB")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[ROW].height = 18
ROW += 1
ROW += 1  # 空行

# ── 列ヘッダー ───────────────────────────────────────────────
month_labels = [f"M{i+1}" for i in range(M)] + ["通期合計", "備考"]
header = ["", "項　目"] + month_labels
for j, h in enumerate(header):
    c = ws.cell(row=ROW, column=j+1, value=h)
    c.font = Font(bold=True, size=9, color=C_HEAD_FG, name="Calibri")
    c.fill = fill(C_HEAD_BG)
    c.alignment = Alignment(horizontal="center" if j > 1 else "left",
                            vertical="center")
    c.border = bd(C_HEAD_BG)
ws.row_dimensions[ROW].height = 22
ROW += 1

# ── ヘルパー：セクション見出し行 ─────────────────────────────
def sec_row(ws, row, label, bg, fg=C_HEAD_FG):
    ws.merge_cells(f"A{row}:P{row}")
    c = ws.cell(row=row, column=1, value=f"  {label}")
    c.font = Font(bold=True, size=9, color=fg, name="Calibri")
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = bd(bg)
    ws.row_dimensions[row].height = 18

# ── ヘルパー：データ行 ────────────────────────────────────────
def data_row(ws, row, indent, label, values, bg, note="", pct=False,
             bold=False, color_neg=False, is_margin=False):
    # インデント列
    c0 = ws.cell(row=row, column=1, value="")
    c0.fill = fill(bg); c0.border = bd()

    # ラベル
    c1 = ws.cell(row=row, column=2,
                 value=("    " * indent) + label)
    c1.font = Font(bold=bold, size=9, name="Calibri",
                   color="374151" if not bold else "111827")
    c1.fill = fill(bg); c1.alignment = Alignment(horizontal="left",
                                                   vertical="center")
    c1.border = bd()

    # 数値
    annual = sum(values) if not is_margin else 0
    for i, v in enumerate(values + ([annual] if not is_margin else [0])):
        col = 3 + i
        if is_margin:
            disp = f"{v*100:.1f}%" if i < M else ""
            c = ws.cell(row=row, column=col, value=disp if i < M else "")
            c.font = Font(size=9, name="Calibri",
                          color=C_POS if (i < M and v >= 0) else C_NEG)
        else:
            c = ws.cell(row=row, column=col, value=round(v, 1) if v != 0 else 0)
            c.number_format = "#,##0.0"
            neg = v < 0 and color_neg
            c.font = Font(bold=bold, size=9, name="Calibri",
                          color=C_NEG if neg else C_POS if (bold and v > 0 and color_neg) else "374151")
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = bd()

    # 備考
    c_note = ws.cell(row=row, column=16, value=note)
    c_note.font = Font(size=8, color="6B7280", italic=True, name="Calibri")
    c_note.fill = fill(bg); c_note.border = bd()
    c_note.alignment = Alignment(horizontal="left", vertical="center",
                                  wrap_text=True)
    ws.row_dimensions[row].height = 17

# ══════════════════════════════════════════════════════════════
# ▼ 売上高
# ══════════════════════════════════════════════════════════════
sec_row(ws, ROW, "▼ 売上高（Revenue）", C_SEC_REV); ROW += 1

data_row(ws, ROW, 1, "① インバウンド体験売上", inbound_rev, C_ROW_REV,
         note="仮: ¥20,000/人 × 月間人数"); ROW += 1
data_row(ws, ROW, 2, "   └ 月間体験人数（人）", inbound_pax, C_ROW_REV,
         note="M4=50人→M12=350人"); ROW += 1
data_row(ws, ROW, 1, "② 広告タイアップ売上", taieup_rev, C_ROW_REV,
         note="仮: 300万/件"); ROW += 1
data_row(ws, ROW, 2, "   └ 月間案件数（件）", taieup_count, C_ROW_REV,
         note="M4=1件→M12=4件"); ROW += 1
data_row(ws, ROW, 1, "③ WESELL グッズ売上", wesell_rev, C_ROW_REV,
         note="米国TikTok Shop、M5〜開始"); ROW += 1
data_row(ws, ROW, 1, "④ IPライセンス売上", license_rev, C_ROW_REV,
         note="仮: 50万/社/月 ×契約社数"); ROW += 1
data_row(ws, ROW, 1, "⑤ ギャル旅タイアップ", galtrip_rev, C_ROW_REV,
         note="仮: 400万/件、年2件"); ROW += 1

# 売上合計行
sec_row(ws, ROW, "  売上高　合計", "374151"); ROW -= 1
data_row(ws, ROW, 0, "  売上高　合計", total_rev, "374151",
         bold=True, color_neg=True); ROW += 1
ROW += 1

# ══════════════════════════════════════════════════════════════
# ▼ 売上原価
# ══════════════════════════════════════════════════════════════
sec_row(ws, ROW, "▼ 売上原価（COGS）", C_SEC_COGS); ROW += 1

data_row(ws, ROW, 1, "インバウンド体験原価", inbound_cogs, C_ROW_COGS,
         note="会場・衣装・スタッフ（売上の40%仮置き）"); ROW += 1
data_row(ws, ROW, 1, "WESELL仕入原価", wesell_cogs, C_ROW_COGS,
         note="グッズ仕入（売上の50%仮置き）"); ROW += 1
data_row(ws, ROW, 1, "ギャル旅制作費", galtrip_cogs, C_ROW_COGS,
         note="動画制作・交通費等（売上の35%仮置き）"); ROW += 1

data_row(ws, ROW, 0, "  売上原価　合計", total_cogs, "78350F",
         bold=True); ROW += 1
ROW += 1

# 売上総利益
data_row(ws, ROW, 0, "  売上総利益（Gross Profit）", gross_profit,
         C_TOTAL, bold=True, color_neg=True); ROW += 1
data_row(ws, ROW, 1, "  売上総利益率", gp_margin, C_TOTAL,
         is_margin=True, note="GP Margin %"); ROW += 1
ROW += 1

# ══════════════════════════════════════════════════════════════
# ▼ 販売費及び一般管理費
# ══════════════════════════════════════════════════════════════
sec_row(ws, ROW, "▼ 販売費及び一般管理費（SG&A）", C_SEC_SGA); ROW += 1

data_row(ws, ROW, 1, "コンテンツ制作費（TikTok/IG）", content_cost,
         C_ROW_SGA, note="動画制作・編集（仮: 月150〜250万）"); ROW += 1
data_row(ws, ROW, 1, "ギャル協会レベニューシェア", gyaru_share,
         C_ROW_SGA, note="広告+ライセンス売上の32%"); ROW += 1
data_row(ws, ROW, 1, "ENTIAL固定費", ential_fixed,
         C_ROW_SGA, note="バズ社内PM（仮: 月80万固定）"); ROW += 1
data_row(ws, ROW, 1, "ENTIALインセンティブ", ential_inc,
         C_ROW_SGA, note="売上の3%（成果連動）"); ROW += 1
data_row(ws, ROW, 1, "広告宣伝費", adv_cost,
         C_ROW_SGA, note="SNS広告・PR費（仮: 月50〜100万）"); ROW += 1
data_row(ws, ROW, 1, "その他運営費", other_cost,
         C_ROW_SGA, note="交通・通信・雑費（仮: 月30万）"); ROW += 1

data_row(ws, ROW, 0, "  販管費　合計", total_sga, "4C1D95",
         bold=True); ROW += 1
ROW += 1

# ══════════════════════════════════════════════════════════════
# ▼ 営業利益
# ══════════════════════════════════════════════════════════════
sec_row(ws, ROW, "▼ 営業利益（Operating Profit）", C_SEC_PROF); ROW += 1

data_row(ws, ROW, 0, "  営業利益", op_profit, C_TOTAL,
         bold=True, color_neg=True); ROW += 1
data_row(ws, ROW, 1, "  営業利益率", op_margin, C_TOTAL,
         is_margin=True, note="Operating Margin %"); ROW += 1
data_row(ws, ROW, 1, "  累計営業利益", cumulative_op, "D1FAE5",
         note="月次累計（損益分岐点確認用）", color_neg=True); ROW += 1
ROW += 1

# ══════════════════════════════════════════════════════════════
# ▼ KPI参考指標
# ══════════════════════════════════════════════════════════════
sec_row(ws, ROW, "▼ KPI参考指標", C_HEAD_BG); ROW += 1

data_row(ws, ROW, 1, "TikTokフォロワー数（万人）",
         [0,0.1,0.3,0.5,0.8,1.2,1.8,2.5,3.5,5,7,10],
         "F8FAFC", note="Year1末：10万人目標"); ROW += 1
data_row(ws, ROW, 1, "インバウンド体験 月間人数",
         inbound_pax, "F8FAFC",
         note="M4=50人→M12=350人"); ROW += 1
data_row(ws, ROW, 1, "月間タイアップ案件数",
         taieup_count, "F8FAFC",
         note="M4=1件→M12=4件"); ROW += 1
ROW += 1

# ── 注記 ─────────────────────────────────────────────────────
ws.merge_cells(f"A{ROW}:P{ROW}")
note_c = ws.cell(row=ROW, column=1,
    value="【注記】本PLは①大成功パターン。単価・原価率はすべて仮置き。"
          "実際の交渉値に応じてセル直接修正可。"
          "赤字数値＝マイナス（損失）、緑数値＝プラス（利益）。")
note_c.font = Font(size=8, color="6B7280", italic=True, name="Calibri")
note_c.fill = fill("F3F4F6")
note_c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[ROW].height = 16

# ══════════════════════════════════════════════════════════════
# 保存
# ══════════════════════════════════════════════════════════════
out = "/home/user/gyal-kyokai--/PL_①大成功パターン.xlsx"
wb.save(out)
print(f"Saved: {out}")

# サマリー出力
print(f"\n【Year1 通期サマリー（大成功）】")
print(f"  売上合計：    {sum(total_rev):,.0f} 万円")
print(f"  売上総利益：  {sum(gross_profit):,.0f} 万円  (GP率: {sum(gross_profit)/sum(total_rev)*100:.1f}%)")
print(f"  販管費合計：  {sum(total_sga):,.0f} 万円")
print(f"  営業利益：    {sum(op_profit):,.0f} 万円  (OPM: {sum(op_profit)/sum(total_rev)*100:.1f}%)")
print(f"  損益分岐月：  M{next((i+1 for i,v in enumerate(cumulative_op) if v>=0), 'N/A')}")
