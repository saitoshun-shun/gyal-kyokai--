"""
事業計画書 - バズ社PL Year1月次 3シナリオ比較
  ① 大成功パターン
  ② 及第点パターン
  ③ 撤退パターン
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════
# カラー定義
# ═══════════════════════════════════════════════════════
NAVY    = "1A1A2E"
NAVY2   = "16213E"
BLUE    = "2E4A7A"
RED_ACC = "C0392B"
ORG_ACC = "D68910"
GREEN   = "1E8449"
LT_BLUE = "D6EAF8"
LT_GRN  = "D5F5E3"
LT_GRN2 = "EAFAF1"
LT_ORG  = "FEF3CE"
LT_RED  = "FADBD8"
LT_GRY  = "F2F3F4"
MID_GRY = "E5E7E9"
WHITE   = "FFFFFF"
GRY_TXT = "7F8C8D"
LT_PUR  = "F4ECF7"

# ═══════════════════════════════════════════════════════
# シナリオ別データ定義
# ═══════════════════════════════════════════════════════

# ─── ① 大成功パターン ─────────────────────────
SUCCESS = {
    "key": "success",
    "sheet": "①大成功",
    "title": "❶ 大成功パターン",
    "subtitle": "Year1でバズ売上8,519万・営業利益1,854万・M6損益分岐",
    "header_color": NAVY,
    "tk_cases": [0, 0, 1, 1, 2, 2, 2, 3, 3, 3, 4, 5],
    "ig_cases": [0, 0, 0, 1, 1, 1, 2, 2, 3, 3, 3, 4],
    "tk_atp_unit": 150,   # 万円/社（統一）
    "ig_atp_unit": 150,
    "ec_units": [0, 0, 50, 120, 250, 450, 700, 1000, 1300, 1700, 2200, 2800],
    "ec_atp":   [0, 0, 5000, 5500, 6000, 6500, 7000, 7200, 7500, 7800, 8000, 8200],
    # 販管費
    "ad_exp": [100, 150, 150, 150, 200, 200, 250, 250, 250, 250, 250, 250],
    "labor":  [100]*12,
    "other":  [50, 50, 50, 60, 60, 60, 70, 70, 80, 80, 90, 90],
    # フォロワー
    "tk_fw":  [500,2000,5000,10000,18000,28000,40000,55000,72000,92000,115000,140000],
    "ig_fw":  [300,1200,3000,6000,12000,20000,30000,43000,58000,75000,95000,120000],
}

# ─── ② 及第点パターン ─────────────────────────
OK_PLAN = {
    "key": "ok",
    "sheet": "②及第点",
    "title": "❷ 及第点パターン",
    "subtitle": "Year1でバズ売上5,510万・営業利益397万・M7損益分岐",
    "header_color": ORG_ACC,
    "tk_cases": [0, 0, 0, 1, 1, 1, 2, 2, 2, 2, 3, 3],
    "ig_cases": [0, 0, 0, 0, 1, 1, 1, 2, 2, 2, 2, 3],
    "tk_atp_unit": 150,
    "ig_atp_unit": 150,
    "ec_units": [0, 0, 30, 70, 130, 250, 400, 600, 800, 1050, 1400, 1800],
    "ec_atp":   [0, 0, 4500, 5000, 5500, 5500, 6000, 6000, 6500, 6500, 7000, 7000],
    "ad_exp": [80, 80, 100, 100, 120, 120, 150, 150, 150, 200, 200, 200],
    "labor":  [100]*12,
    "other":  [50, 50, 50, 50, 60, 60, 60, 70, 70, 80, 80, 80],
    "tk_fw":  [300,1000,2500,5000,9000,15000,22000,30000,40000,52000,66000,82000],
    "ig_fw":  [200, 600,1500,3000,6000,10000,15000,21000,28000,36000,46000,58000],
}

# ─── ③ 撤退パターン ──────────────────────────
RETREAT = {
    "key": "retreat",
    "sheet": "③撤退",
    "title": "❸ 撤退パターン",
    "subtitle": "Year1でバズ売上715万・営業損失2,435万・M9末で撤退決定",
    "header_color": RED_ACC,
    "tk_cases": [0, 0, 0, 0, 1, 0, 1, 0, 1, 1, 0, 0],
    "ig_cases": [0, 0, 0, 0, 0, 1, 0, 0, 1, 0, 0, 0],
    "tk_atp_unit": 100,
    "ig_atp_unit": 100,
    "ec_units": [0, 0, 10, 20, 40, 80, 100, 130, 150, 180, 200, 250],
    "ec_atp":   [0, 0, 4000, 4000, 4500, 4500, 5000, 5000, 5000, 5000, 5000, 5000],
    "ad_exp": [80, 80, 80, 80, 100, 100, 120, 120, 100, 80, 50, 50],
    "labor":  [100]*12,
    "other":  [50, 50, 50, 50, 50, 50, 50, 50, 50, 50, 50, 50],
    "tk_fw":  [200, 500,1000,1500,2000,2800,3500,4000,4500,4800,5000,5000],
    "ig_fw":  [100, 300, 600, 900,1200,1600,2000,2300,2500,2700,2800,2800],
}

SCENARIOS = [SUCCESS, OK_PLAN, RETREAT]

# ═══════════════════════════════════════════════════════
# 各シナリオの計算（派生値）
# ═══════════════════════════════════════════════════════
def compute(sc):
    tk_tie = [c * sc["tk_atp_unit"] for c in sc["tk_cases"]]
    ig_tie = [c * sc["ig_atp_unit"] for c in sc["ig_cases"]]
    tie_total = [t + i for t, i in zip(tk_tie, ig_tie)]

    # EC：ギャル協会がショップオーナー → バズはアフィリエイトコミッション（×20%）のみ
    ec_gross      = [round(u * a / 10000) for u, a in zip(sc["ec_units"], sc["ec_atp"])]
    ec_commission = [round(e * 0.20) for e in ec_gross]

    # バズ売上 = タイアップ + ECアフィリエイトコミッション
    rev_total = [t + ec for t, ec in zip(tie_total, ec_commission)]

    # ギャル協会出演・監修料: 件数×25万（月最低保証10万）
    gal_cases   = [tk + ig for tk, ig in zip(sc["tk_cases"], sc["ig_cases"])]
    tie_license = [max(c * 25, 10) for c in gal_cases]
    tie_prod    = [round(t * 0.10) for t in tie_total]
    tie_cogs    = [l + p for l, p in zip(tie_license, tie_prod)]
    tie_gross   = [t - cg for t, cg in zip(tie_total, tie_cogs)]

    # EC原価はバズに発生しない（ギャル協会がショップオーナー）
    cogs_total  = tie_cogs
    gross_total = [r - c for r, c in zip(rev_total, cogs_total)]

    # ENTIAL成功報酬: TIE×5%（体験事業元締めがENTIALの主収益、EC成功報酬は廃止）
    ential_fee = [round(t * 0.05) for t in tie_total]

    sga = [a + l + o + en
           for a, l, o, en in zip(sc["ad_exp"], sc["labor"], sc["other"], ential_fee)]
    op_profit = [g - s for g, s in zip(gross_total, sga)]
    bep_month = next((i for i, v in enumerate(op_profit) if v > 0), None)

    gp_rate  = [round(g/r*100,1) if r>0 else 0.0 for g,r in zip(gross_total, rev_total)]
    tie_gp_r = [round(tg/t*100,1) if t>0 else 0.0 for tg,t in zip(tie_gross, tie_total)]
    op_rate  = [round(o/r*100,1) if r>0 else 0.0 for o,r in zip(op_profit, rev_total)]

    return dict(
        tk_tie=tk_tie, ig_tie=ig_tie, tie_total=tie_total,
        tk_atp=[sc["tk_atp_unit"] if c>0 else 0 for c in sc["tk_cases"]],
        ig_atp=[sc["ig_atp_unit"] if c>0 else 0 for c in sc["ig_cases"]],
        ec_gross=ec_gross, ec_commission=ec_commission,
        rev_total=rev_total,
        gal_cases=gal_cases,
        tie_license=tie_license, tie_prod=tie_prod, tie_cogs=tie_cogs,
        tie_gross=tie_gross, tie_gp_r=tie_gp_r,
        cogs_total=cogs_total, gross_total=gross_total, gp_rate=gp_rate,
        ential_fee=ential_fee,
        sga=sga, op_profit=op_profit, op_rate=op_rate, bep_month=bep_month,
    )

# 各シナリオの計算結果を格納
for sc in SCENARIOS:
    sc["c"] = compute(sc)

# ═══════════════════════════════════════════════════════
# 共通ヘルパー
# ═══════════════════════════════════════════════════════
MONTHS = [f"M{i}" for i in range(1, 13)]
NCOL_ITEM, NCOL_M1, NCOL_TOTAL = 1, 2, 14

def set_cell(ws, row, col, val="", bold=False, sz=10, fg="222222", bg=None,
             ha="center", italic=False, fmt=None, indent=0):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(name="Meiryo UI", bold=bold, size=sz,
                     color=fg, italic=italic)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=ha, vertical="center", indent=indent)
    if fmt:
        cell.number_format = fmt
    return cell

def fill_empty_row(ws, row, col_start, col_end, bg):
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=bg)

STYLE = {
    "sec":       dict(bold=True,  sz=9,  fg="1A1A2E", bg=MID_GRY,  ha="left"),
    "normal":    dict(bold=False, sz=9,  fg="333333", bg=WHITE,     ha="left"),
    "kpi_sub":   dict(bold=False, sz=8,  fg="888888", bg="FAFAFA",  ha="left"),
    "sub":       dict(bold=True,  sz=9,  fg="1A3A6B", bg=LT_BLUE,   ha="left"),
    "total_rev": dict(bold=True,  sz=10, fg="1A3A6B", bg=LT_BLUE,   ha="left"),
    "total_cost":dict(bold=True,  sz=9,  fg="555555", bg=LT_GRY,    ha="left"),
    "tie_gross": dict(bold=True,  sz=10, fg="4A235A", bg=LT_PUR,    ha="left"),
    "ec_gross":  dict(bold=True,  sz=10, fg="145A32", bg=LT_GRN2,   ha="left"),
    "gross":     dict(bold=True,  sz=10, fg="145A32", bg=LT_GRN,    ha="left"),
    "rate":      dict(bold=False, sz=9,  fg=GRY_TXT,  bg=LT_GRY,    ha="left"),
    "op":        dict(bold=True,  sz=11, fg=WHITE,    bg=NAVY,      ha="left"),
    "blank":     dict(bold=False, sz=6,  fg=WHITE,    bg=WHITE,     ha="left"),
}

# ═══════════════════════════════════════════════════════
# PLシート生成関数
# ═══════════════════════════════════════════════════════
def build_pl_sheet(ws, sc):
    c_data = sc["c"]
    header_col = sc["header_color"]
    bep = c_data["bep_month"]

    # 列幅設定
    ws.column_dimensions["A"].width = 36
    for col in range(2, 15):
        ws.column_dimensions[get_column_letter(col)].width = 10

    ROW = 1
    # タイトル
    ws.row_dimensions[ROW].height = 38
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=14)
    cell = ws.cell(row=ROW, column=1,
                   value=f"{sc['title']}  ／  Year1 月次事業計画（バズ社PL）")
    cell.font = Font(name="Meiryo UI", bold=True, size=16, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=header_col)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ROW += 1
    ws.row_dimensions[ROW].height = 16
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=14)
    cell = ws.cell(row=ROW, column=1, value=sc["subtitle"] + "　単位：万円")
    cell.font = Font(name="Meiryo UI", size=9, color="CCCCCC")
    cell.fill = PatternFill("solid", fgColor=NAVY2)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # ヘッダー行
    ROW += 1
    ws.row_dimensions[ROW].height = 26
    set_cell(ws, ROW, 1, "項目", bold=True, sz=10, fg=WHITE, bg=BLUE)
    for i, m in enumerate(MONTHS):
        set_cell(ws, ROW, 2+i, m, bold=True, sz=10, fg=WHITE, bg=BLUE)
    set_cell(ws, ROW, 14, "Year1合計", bold=True, sz=10, fg=WHITE, bg=BLUE)
    HDR_ROW = ROW

    # KPIセクション
    ROW += 1
    ws.row_dimensions[ROW].height = 22
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=14)
    kpi_cell = ws.cell(row=ROW, column=1, value="▌ 重要KPI（月次推移）")
    kpi_cell.font = Font(name="Meiryo UI", bold=True, size=11, color=WHITE)
    kpi_cell.fill = PatternFill("solid", fgColor=BLUE)
    kpi_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    kpi_rows = [
        ("TikTokフォロワー数（累計）", sc["tk_fw"], "#,##0", True, "last"),
        ("IGフォロワー数（累計）",     sc["ig_fw"], "#,##0", True, "last"),
        ("TKタイアップ受注社数（月次）", sc["tk_cases"], "#,##0", True, "sum"),
        ("TKタイアップ平均単価（万円/社）", c_data["tk_atp"], "#,##0", False, "avg_nz"),
        ("IGタイアップ受注社数（月次）", sc["ig_cases"], "#,##0", True, "sum"),
        ("IGタイアップ平均単価（万円/社）", c_data["ig_atp"], "#,##0", False, "avg_nz"),
        ("EC注文件数（TikTok Shop）",  sc["ec_units"], "#,##0", True, "sum"),
        ("EC平均注文単価（円）",       sc["ec_atp"], "#,##0", False, "avg_nz"),
    ]
    for ki, (label, vals, fmt, is_kpi, agg) in enumerate(kpi_rows):
        ROW += 1
        ws.row_dimensions[ROW].height = 17
        row_bg = LT_GRY if ki % 2 == 0 else WHITE
        set_cell(ws, ROW, 1, label, bold=is_kpi, sz=9,
                 fg=BLUE if is_kpi else "333333", bg=row_bg, ha="left", indent=1)
        for j, v in enumerate(vals):
            set_cell(ws, ROW, 2+j, v if v != 0 else "-",
                     bold=is_kpi, sz=9,
                     fg=BLUE if is_kpi else "444444", bg=row_bg,
                     fmt=fmt if v != 0 else None)
        # 集計
        non_z = [v for v in vals if v != 0]
        if agg == "last":    agg_v = vals[-1]
        elif agg == "sum":   agg_v = sum(vals)
        elif agg == "avg_nz":agg_v = round(sum(non_z)/len(non_z)) if non_z else 0
        set_cell(ws, ROW, 14, agg_v if agg_v != 0 else "-",
                 bold=True, sz=9,
                 fg=BLUE if is_kpi else "333333", bg=row_bg,
                 fmt=fmt if agg_v != 0 else None)

    # PLセクション ヘッダー
    ROW += 1
    ws.row_dimensions[ROW].height = 22
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=14)
    pl_cell = ws.cell(row=ROW, column=1, value="▌ 損益計算書（P/L）　単位：万円")
    pl_cell.font = Font(name="Meiryo UI", bold=True, size=11, color=WHITE)
    pl_cell.fill = PatternFill("solid", fgColor=NAVY)
    pl_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # PL行定義
    pl_rows = [
        ("【売上高】",                                       None,                    "sec",     None,     "#,##0"),
        ("  ① TikTokメディアタイアップ",                    c_data["tk_tie"],        "normal",  "sum",    "#,##0"),
        ("     受注社数（社）",                              sc["tk_cases"],          "kpi_sub", "sum",    "#,##0"),
        ("     平均単価（万円/社）",                         c_data["tk_atp"],        "kpi_sub", "avg_nz", "#,##0"),
        ("  ② IGメディアタイアップ",                        c_data["ig_tie"],        "normal",  "sum",    "#,##0"),
        ("     受注社数（社）",                              sc["ig_cases"],          "kpi_sub", "sum",    "#,##0"),
        ("     平均単価（万円/社）",                         c_data["ig_atp"],        "kpi_sub", "avg_nz", "#,##0"),
        ("  タイアップ売上合計",                             c_data["tie_total"],     "sub",     "sum",    "#,##0"),
        ("  ③ ECアフィリエイトコミッション（×20%）",        c_data["ec_commission"], "normal",  "sum",    "#,##0"),
        ("     EC実売上（ギャル協会側・参考値）",            c_data["ec_gross"],      "kpi_sub", "sum",    "#,##0"),
        ("     注文件数（件）",                              sc["ec_units"],          "kpi_sub", "sum",    "#,##0"),
        ("     平均注文単価（円）",                          sc["ec_atp"],            "kpi_sub", "avg_nz", "#,##0"),
        ("売上合計",                                        c_data["rev_total"],     "total_rev","sum",   "#,##0"),
        ("", None, "blank", None, None),
        ("【タイアップ原価】",                               None,                    "sec",     None,     "#,##0"),
        ("  ギャル協会出演・監修料（件数×25万・最低保証10万）", c_data["tie_license"], "normal", "sum",   "#,##0"),
        ("     タイアップ件数（社）",                        c_data["gal_cases"],     "kpi_sub", "sum",   "#,##0"),
        ("  コンテンツ制作費（×10%）",                      c_data["tie_prod"],      "normal",  "sum",   "#,##0"),
        ("  タイアップ原価合計",                             c_data["tie_cogs"],      "sub",     "sum",   "#,##0"),
        ("  タイアップ粗利",                                 c_data["tie_gross"],     "tie_gross","sum",  "#,##0"),
        ("  タイアップ粗利率",                               c_data["tie_gp_r"],      "rate",    "avg_nz", None),
        ("", None, "blank", None, None),
        ("  ※ECコミッション（原価ゼロ）→ ギャル協会がショップオーナー", None, "kpi_sub", None, None),
        ("", None, "blank", None, None),
        ("売上原価合計",                                    c_data["cogs_total"],    "total_cost","sum",  "#,##0"),
        ("粗利益合計",                                      c_data["gross_total"],   "gross",   "sum",   "#,##0"),
        ("粗利率（合計）",                                  c_data["gp_rate"],       "rate",    "avg_nz", None),
        ("", None, "blank", None, None),
        ("【販売管理費】",                                   None,                    "sec",     None,    "#,##0"),
        ("  広告費（TK/IG広告運用）",                       sc["ad_exp"],            "normal",  "sum",   "#,##0"),
        ("  人件費（バズ社員）",                             sc["labor"],             "normal",  "sum",   "#,##0"),
        ("  その他販管費",                                   sc["other"],             "normal",  "sum",   "#,##0"),
        ("  ENTIAL業務委託費（TIE×5%）",                   c_data["ential_fee"],    "normal",  "sum",   "#,##0"),
        ("販売管理費合計",                                   c_data["sga"],           "total_cost","sum", "#,##0"),
        ("", None, "blank", None, None),
        ("営業利益",                                        c_data["op_profit"],     "op",      "sum",   "#,##0"),
        ("営業利益率",                                      c_data["op_rate"],       "rate",    "avg_nz", None),
    ]

    for (label, vals, style, agg_method, num_fmt) in pl_rows:
        ROW += 1
        ws.row_dimensions[ROW].height = (6 if style=="blank"
                                         else 15 if style=="kpi_sub" else 18)
        st = STYLE[style]
        cell = ws.cell(row=ROW, column=1, value=label)
        cell.font = Font(name="Meiryo UI", bold=st["bold"], size=st["sz"], color=st["fg"])
        cell.fill = PatternFill("solid", fgColor=st["bg"])
        cell.alignment = Alignment(horizontal=st["ha"], vertical="center", indent=1)

        if vals is None:
            fill_empty_row(ws, ROW, 2, 14, st["bg"])
            continue

        is_rate = (style == "rate")
        cell_fmt = ('0.0"%"' if is_rate else (num_fmt or "#,##0"))

        for j, v in enumerate(vals):
            col = 2 + j
            disp_val = "-" if (v == 0 and style in ("kpi_sub",)) else v
            cell2 = ws.cell(row=ROW, column=col, value=disp_val)
            if style == "op" and j == bep:
                cell2.fill = PatternFill("solid", fgColor=GREEN)
                cell2.font = Font(name="Meiryo UI", bold=True, size=st["sz"], color=WHITE)
            else:
                cell2.fill = PatternFill("solid", fgColor=st["bg"])
                fg_v = (WHITE if style=="op" and v>=0
                        else "FFCCCC" if style=="op" and v<0
                        else "145A32" if (style in ("gross","ec_gross","tie_gross") and v>=0)
                        else RED_ACC if (style in ("gross","ec_gross","tie_gross") and v<0)
                        else st["fg"])
                cell2.font = Font(name="Meiryo UI", bold=st["bold"], size=st["sz"], color=fg_v)
            cell2.alignment = Alignment(horizontal="center", vertical="center")
            if disp_val != "-":
                cell2.number_format = cell_fmt

        # 合計列
        non_z2 = [v for v in vals if v != 0]
        fmt2 = '0.0"%"' if is_rate else (num_fmt or "#,##0")
        if agg_method == "avg_nz":
            agg_v = round(sum(non_z2)/len(non_z2), 1) if non_z2 else 0.0
        else:
            agg_v = sum(vals)

        disp_agg = "-" if (agg_v == 0 and style == "kpi_sub") else agg_v
        cell3 = ws.cell(row=ROW, column=14, value=disp_agg)
        fg3 = (WHITE if style == "op" else
               ("145A32" if (style in ("gross","ec_gross","tie_gross") and agg_v >= 0)
                else RED_ACC if (style in ("gross","ec_gross","tie_gross") and agg_v < 0)
                else st["fg"]))
        bg3 = (GREEN if (style == "op" and agg_v >= 0)
               else RED_ACC if (style == "op" and agg_v < 0)
               else st["bg"])
        cell3.font = Font(name="Meiryo UI", bold=True, size=st["sz"], color=fg3)
        cell3.fill = PatternFill("solid", fgColor=bg3)
        cell3.alignment = Alignment(horizontal="center", vertical="center")
        if disp_agg != "-":
            cell3.number_format = fmt2

    # BEP注記
    ROW += 1
    ws.row_dimensions[ROW].height = 18
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=14)
    bep_label = f"M{bep+1}" if bep is not None else "Year1内で達成せず"
    total_op = sum(c_data["op_profit"])
    total_rev = sum(c_data["rev_total"])
    note = (f"★ 損益分岐：{bep_label}　／　"
            f"Year1通期営業利益：{total_op:,}万円　／　"
            f"Year1売上：{total_rev:,}万円")
    ncell = ws.cell(row=ROW, column=1, value=note)
    ncell.font = Font(name="Meiryo UI", size=9, bold=True,
                      color=NAVY if total_op >= 0 else RED_ACC)
    ncell.fill = PatternFill("solid", fgColor=LT_GRN if total_op >= 0 else LT_RED)
    ncell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # フッター
    ROW += 1
    ws.row_dimensions[ROW].height = 14
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=14)
    fcell = ws.cell(row=ROW, column=1,
        value="【注記】単位：万円 ／ ギャル協会：出演・監修料=件数×25万（最低保証10万）、EC物販はギャル協会がショップオーナー（バズに原価なし） ／ "
              "バズEC収益=アフィリエイトコミッション（EC実売上×20%） ／ 制作費：TIE×10% ／ ENTIAL：TIE成功報酬×5%（EC運用除外）")
    fcell.font = Font(name="Meiryo UI", size=8, color=GRY_TXT, italic=True)
    fcell.fill = PatternFill("solid", fgColor=LT_GRY)
    fcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # 枠線
    thin = Side(style="thin", color="DDDDDD")
    for row in ws.iter_rows(min_row=HDR_ROW, max_row=ROW, min_col=1, max_col=14):
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws.freeze_panes = "B4"


# ═══════════════════════════════════════════════════════
# 比較・撤退基準シート生成関数
# ═══════════════════════════════════════════════════════
def build_comparison_sheet(ws):
    # 列幅
    ws.column_dimensions["A"].width = 32
    for col in range(2, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18

    ROW = 1
    # タイトル
    ws.row_dimensions[ROW].height = 40
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=5)
    cell = ws.cell(row=ROW, column=1,
                   value="3シナリオ比較 ＆ 撤退判断基準")
    cell.font = Font(name="Meiryo UI", bold=True, size=18, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ROW += 1
    ws.row_dimensions[ROW].height = 16
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=5)
    cell = ws.cell(row=ROW, column=1,
                   value="ギャル×インバウンド共同事業（バズ社PL）　Year1・月次単位")
    cell.font = Font(name="Meiryo UI", size=9, color="CCCCCC")
    cell.fill = PatternFill("solid", fgColor=NAVY2)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # ─── Year1通期サマリー比較 ───
    ROW += 2
    ws.row_dimensions[ROW].height = 24
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=5)
    cell = ws.cell(row=ROW, column=1, value="▌ Year1通期サマリー（3シナリオ比較）")
    cell.font = Font(name="Meiryo UI", bold=True, size=12, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ヘッダー
    ROW += 1
    ws.row_dimensions[ROW].height = 24
    headers = ["項目", "①大成功", "②及第点", "③撤退"]
    for i, h in enumerate(headers):
        color = [BLUE, NAVY, ORG_ACC, RED_ACC][i]
        set_cell(ws, ROW, 1+i, h, bold=True, sz=11, fg=WHITE, bg=color,
                 ha="center" if i > 0 else "left", indent=1 if i == 0 else 0)

    # サマリー行
    rows_data = [
        ("売上合計（万円）",              [sum(sc["c"]["rev_total"]) for sc in SCENARIOS],      "#,##0"),
        ("  ├ タイアップ売上",           [sum(sc["c"]["tie_total"]) for sc in SCENARIOS],      "#,##0"),
        ("  └ ECアフィリエイトコミッション（×20%）",[sum(sc["c"]["ec_commission"]) for sc in SCENARIOS], "#,##0"),
        ("  ※EC実売上（ギャル協会側・参考）",[sum(sc["c"]["ec_gross"]) for sc in SCENARIOS],   "#,##0"),
        ("粗利合計（万円）",              [sum(sc["c"]["gross_total"]) for sc in SCENARIOS],    "#,##0"),
        ("粗利率",                        [round(sum(sc["c"]["gross_total"])/sum(sc["c"]["rev_total"])*100,1) if sum(sc["c"]["rev_total"])>0 else 0 for sc in SCENARIOS], '0.0"%"'),
        ("販管費合計（万円）",            [sum(sc["c"]["sga"]) for sc in SCENARIOS],            "#,##0"),
        ("  うちENTIAL成功報酬（TIE×5%）",[sum(sc["c"]["ential_fee"]) for sc in SCENARIOS],   "#,##0"),
        ("営業利益（万円）",              [sum(sc["c"]["op_profit"]) for sc in SCENARIOS],      "#,##0"),
        ("営業利益率",                    [round(sum(sc["c"]["op_profit"])/sum(sc["c"]["rev_total"])*100,1) if sum(sc["c"]["rev_total"])>0 else 0 for sc in SCENARIOS], '0.0"%"'),
        ("損益分岐点",                    [f"M{sc['c']['bep_month']+1}" if sc['c']['bep_month'] is not None else "達成せず" for sc in SCENARIOS], None),
        ("Year1末 TKフォロワー",          [sc["tk_fw"][-1] for sc in SCENARIOS],               "#,##0"),
        ("Year1末 IGフォロワー",          [sc["ig_fw"][-1] for sc in SCENARIOS],               "#,##0"),
        ("Year1 タイアップ累計社数",      [sum(sc["tk_cases"])+sum(sc["ig_cases"]) for sc in SCENARIOS], "#,##0"),
        ("Year1 EC累計注文件数（参考）",  [sum(sc["ec_units"]) for sc in SCENARIOS],            "#,##0"),
    ]
    for ri, (label, vals, fmt) in enumerate(rows_data):
        ROW += 1
        ws.row_dimensions[ROW].height = 22
        is_highlight = label in ("売上合計（万円）", "営業利益（万円）", "損益分岐点")
        bg = LT_BLUE if is_highlight else (LT_GRY if ri % 2 == 0 else WHITE)
        set_cell(ws, ROW, 1, label, bold=is_highlight, sz=10,
                 fg="1A3A6B" if is_highlight else "333333",
                 bg=bg, ha="left", indent=1)
        for i, v in enumerate(vals):
            col = 2 + i
            # 営業利益は色付け
            fg_c = "333333"
            if label == "営業利益（万円）":
                fg_c = GREEN if v >= 0 else RED_ACC
            elif label == "損益分岐点":
                fg_c = GREEN if v != "達成せず" else RED_ACC
            set_cell(ws, ROW, col, v, bold=is_highlight, sz=10,
                     fg=fg_c, bg=bg, fmt=fmt if isinstance(v, (int, float)) else None)

    # ─── 各社受取シミュレーション ───
    ROW += 2
    ws.row_dimensions[ROW].height = 24
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=5)
    cell = ws.cell(row=ROW, column=1, value="▌ 各社Year1受取シミュレーション（単位：万円）")
    cell.font = Font(name="Meiryo UI", bold=True, size=12, color=WHITE)
    cell.fill = PatternFill("solid", fgColor="4A235A")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ROW += 1
    ws.row_dimensions[ROW].height = 22
    for i, h in enumerate(["受取者・項目", "①大成功", "②及第点", "③撤退"]):
        color = ["4A235A", NAVY, ORG_ACC, RED_ACC][i]
        set_cell(ws, ROW, 1+i, h, bold=True, sz=10, fg=WHITE, bg=color,
                 ha="center" if i > 0 else "left", indent=1 if i == 0 else 0)

    payout_rows = [
        ("【ギャル協会 受取合計】",
         [sum(sc["c"]["tie_license"]) for sc in SCENARIOS],
         "#,##0", True),
        ("  出演・監修料（件数×25万・最低保証10万）",
         [sum(sc["c"]["tie_license"]) for sc in SCENARIOS], "#,##0", False),
        ("  ※EC物販収益はギャル協会のショップ利益（バズPL外）",
         ["-"] * len(SCENARIOS), None, False),
        ("【ENTIAL 受取合計】（TIE成功報酬のみ）",
         [sum(sc["c"]["ential_fee"]) for sc in SCENARIOS], "#,##0", True),
        ("  タイアップ成功報酬（TIE×5%）",
         [sum(sc["c"]["ential_fee"]) for sc in SCENARIOS], "#,##0", False),
        ("  ※体験事業元締め収益は別途（バズPL外）",
         ["-"] * len(SCENARIOS), None, False),
        ("【バズ ECコミッション収益】",
         [sum(sc["c"]["ec_commission"]) for sc in SCENARIOS], "#,##0", True),
        ("【バズ 営業利益（手残り）】",
         [sum(sc["c"]["op_profit"]) for sc in SCENARIOS], "#,##0", True),
    ]
    for pi, (label, vals, fmt, bold) in enumerate(payout_rows):
        ROW += 1
        ws.row_dimensions[ROW].height = 22
        bg = LT_PUR if bold else (LT_GRY if pi % 2 == 0 else WHITE)
        set_cell(ws, ROW, 1, label, bold=bold, sz=10,
                 fg="4A235A" if bold else "333333", bg=bg, ha="left", indent=1)
        for i, v in enumerate(vals):
            col = 2 + i
            fg_c = "333333"
            if label == "【バズ 営業利益（手残り）】":
                fg_c = GREEN if v >= 0 else RED_ACC
            set_cell(ws, ROW, col, v, bold=bold, sz=10,
                     fg=fg_c, bg=bg, fmt=fmt)
        set_cell(ws, ROW, 5, "", bg=bg)

    # ─── 撤退判断のKPIゲート ───
    ROW += 2
    ws.row_dimensions[ROW].height = 24
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=5)
    cell = ws.cell(row=ROW, column=1, value="▌ 撤退判断のKPIゲート（3段階チェック）")
    cell.font = Font(name="Meiryo UI", bold=True, size=12, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=RED_ACC)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    gates = [
        # (タイミング, チェック項目, 撤退基準（下回ると危険）, アクション)
        ("M3末", "TKフォロワー", "< 2,000人", "→ 改善施策を検討"),
        ("M3末", "IGフォロワー", "< 1,000人", "→ 改善施策を検討"),
        ("M6末", "タイアップ累計受注", "< 3社", "→ イエローカード、営業強化"),
        ("M6末", "EC累計注文件数",   "< 200件", "→ イエローカード、施策見直し"),
        ("M6末", "単月営業利益",     "< -300万円", "→ コスト圧縮検討"),
        ("M9末", "月次タイアップ受注", "< 2社", "→ ★撤退決定"),
        ("M9末", "月次EC売上",         "< 300万円", "→ ★撤退決定"),
        ("M9末", "単月営業利益",       "< -200万円継続", "→ ★撤退決定"),
        ("M10〜", "撤退プロセス実行",  "新規受注停止", "既存案件のみクロージング"),
        ("M12末", "事業終了",          "完全撤退",     "Year1末で事業停止"),
    ]
    ROW += 1
    ws.row_dimensions[ROW].height = 22
    for i, h in enumerate(["タイミング","チェック項目","撤退基準","アクション"]):
        set_cell(ws, ROW, 1+i, h, bold=True, sz=10, fg=WHITE, bg=BLUE, ha="center")
    # 5列目は空白
    set_cell(ws, ROW, 5, "", bg=BLUE)

    for gi, (timing, item, criteria, action) in enumerate(gates):
        ROW += 1
        ws.row_dimensions[ROW].height = 20
        is_red = "★撤退" in action or "撤退プロセス" in item or "事業終了" in item
        bg = LT_RED if is_red else (LT_GRY if gi % 2 == 0 else WHITE)
        for i, v in enumerate([timing, item, criteria, action]):
            fg_c = RED_ACC if is_red else "333333"
            set_cell(ws, ROW, 1+i, v, bold=is_red, sz=9,
                     fg=fg_c, bg=bg, ha="left", indent=1)
        set_cell(ws, ROW, 5, "", bg=bg)

    # ─── ①大成功の成立条件 ───
    ROW += 2
    ws.row_dimensions[ROW].height = 24
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=5)
    cell = ws.cell(row=ROW, column=1, value="▌ ①大成功パターンに着地するための条件")
    cell.font = Font(name="Meiryo UI", bold=True, size=12, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=GREEN)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    success_conds = [
        ("M3時点",    "TKフォロワー 5,000人以上、IGフォロワー 3,000人以上"),
        ("M3時点",    "タイアップ初回受注（1件以上）を獲得"),
        ("M6時点",    "TKフォロワー 28,000人以上、IGフォロワー 20,000人以上"),
        ("M6時点",    "タイアップ月次受注 4件以上（TK+IG合算）"),
        ("M6時点",    "EC月次注文 450件以上・月次売上 290万円以上"),
        ("M9時点",    "タイアップ単価 150万円/社を維持"),
        ("M9時点",    "EC月次売上 975万円以上"),
        ("前提条件",  "ENTIALがバズ社内PMとして常駐し、営業・オペをすべて巻き取る"),
        ("前提条件",  "うさたにパイセン出演コンテンツを安定供給できる"),
        ("前提条件",  "WESELL（TikTok Shop）が安定運用できるインフラ整備"),
    ]
    ROW += 1
    ws.row_dimensions[ROW].height = 22
    for i, h in enumerate(["タイミング","条件"]):
        set_cell(ws, ROW, 1+i, h, bold=True, sz=10, fg=WHITE, bg=BLUE,
                 ha="center" if i > 0 else "left", indent=1 if i == 0 else 0)
    ws.merge_cells(start_row=ROW, start_column=2, end_row=ROW, end_column=5)

    for ci, (timing, cond) in enumerate(success_conds):
        ROW += 1
        ws.row_dimensions[ROW].height = 20
        bg = LT_GRN2 if ci % 2 == 0 else WHITE
        is_premise = timing == "前提条件"
        set_cell(ws, ROW, 1, timing, bold=True, sz=9,
                 fg=GREEN if is_premise else "333333",
                 bg=bg, ha="left", indent=1)
        ws.merge_cells(start_row=ROW, start_column=2, end_row=ROW, end_column=5)
        set_cell(ws, ROW, 2, cond, bold=False, sz=9, fg="333333", bg=bg, ha="left", indent=1)

    # 枠線
    thin = Side(style="thin", color="DDDDDD")
    for row in ws.iter_rows(min_row=1, max_row=ROW, min_col=1, max_col=5):
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


# ═══════════════════════════════════════════════════════
# 内訳シート生成関数（バズの支払先ごとに整理）
# ═══════════════════════════════════════════════════════
def build_breakdown_sheet(ws, sc):
    """バズが誰にいくら払うか、月次で確認できる支払い内訳シート"""
    c_data = sc["c"]
    header_col = sc["header_color"]
    bep = c_data["bep_month"]

    ws.column_dimensions["A"].width = 38
    for col in range(2, 15):
        ws.column_dimensions[get_column_letter(col)].width = 10

    ROW = 1
    ws.row_dimensions[ROW].height = 38
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=14)
    cell = ws.cell(row=ROW, column=1,
                   value=f"{sc['title']}  ／  バズ 収益・支払い内訳（月次）")
    cell.font = Font(name="Meiryo UI", bold=True, size=16, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=header_col)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ROW += 1
    ws.row_dimensions[ROW].height = 16
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=14)
    cell = ws.cell(row=ROW, column=1,
                   value="バズが誰にいくら支払うか月次で確認できます　単位：万円")
    cell.font = Font(name="Meiryo UI", size=9, color="CCCCCC")
    cell.fill = PatternFill("solid", fgColor=NAVY2)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ROW += 1
    ws.row_dimensions[ROW].height = 26
    set_cell(ws, ROW, 1, "項目", bold=True, sz=10, fg=WHITE, bg=BLUE)
    for i, m in enumerate(MONTHS):
        set_cell(ws, ROW, 2+i, m, bold=True, sz=10, fg=WHITE, bg=BLUE)
    set_cell(ws, ROW, 14, "Year1合計", bold=True, sz=10, fg=WHITE, bg=BLUE)
    HDR_ROW = ROW

    rows = [
        # ── 収益 ──────────────────────────────────
        ("【バズ 収益内訳】",                                     None,                    "sec",      None,     "#,##0"),
        ("  ① TKタイアップ売上",                                  c_data["tk_tie"],        "normal",   "sum",    "#,##0"),
        ("     受注社数（社）",                                   sc["tk_cases"],          "kpi_sub",  "sum",    "#,##0"),
        ("  ② IGタイアップ売上",                                  c_data["ig_tie"],        "normal",   "sum",    "#,##0"),
        ("     受注社数（社）",                                   sc["ig_cases"],          "kpi_sub",  "sum",    "#,##0"),
        ("  タイアップ売上計",                                    c_data["tie_total"],     "sub",      "sum",    "#,##0"),
        ("  ③ ECアフィリエイトコミッション（×20%）",              c_data["ec_commission"], "normal",   "sum",    "#,##0"),
        ("     ※EC実売上（ギャル協会ショップ・参考）",            c_data["ec_gross"],      "kpi_sub",  "sum",    "#,##0"),
        ("バズ 売上合計",                                         c_data["rev_total"],     "total_rev","sum",    "#,##0"),
        ("", None, "blank", None, None),

        # ── バズ→ギャル協会 ──────────────────────
        ("【バズ → ギャル協会 への支払い】",                       None,                    "sec",      None,     "#,##0"),
        ("  出演・監修料（件数×25万・最低保証10万）",             c_data["tie_license"],   "normal",   "sum",    "#,##0"),
        ("     タイアップ件数（社）",                             c_data["gal_cases"],     "kpi_sub",  "sum",    "#,##0"),
        ("", None, "blank", None, None),

        # ── バズ→制作費 ──────────────────────────
        ("【バズ → コンテンツ制作費】",                            None,                    "sec",      None,     "#,##0"),
        ("  制作費（タイアップ売上×10%）",                       c_data["tie_prod"],      "normal",   "sum",    "#,##0"),
        ("", None, "blank", None, None),

        # ── タイアップ原価まとめ ──────────────────
        ("タイアップ原価合計（出演料＋制作費）",                  c_data["tie_cogs"],      "total_cost","sum",   "#,##0"),
        ("タイアップ粗利",                                        c_data["tie_gross"],     "tie_gross","sum",    "#,##0"),
        ("タイアップ粗利率",                                      c_data["tie_gp_r"],      "rate",     "avg_nz", None),
        ("", None, "blank", None, None),
        ("  ※ ECコミッション：バズの原価ゼロ（ギャル協会がショップオーナー）",
                                                                  None,                    "kpi_sub",  None,     None),
        ("", None, "blank", None, None),
        ("バズ 粗利合計",                                         c_data["gross_total"],   "gross",    "sum",    "#,##0"),
        ("粗利率（合計）",                                        c_data["gp_rate"],       "rate",     "avg_nz", None),
        ("", None, "blank", None, None),

        # ── バズ→広告・運営費 ────────────────────
        ("【バズ → 広告・運営費】",                                None,                    "sec",      None,     "#,##0"),
        ("  広告費（TK/IG広告運用）",                            sc["ad_exp"],            "normal",   "sum",    "#,##0"),
        ("  人件費（バズ社員）",                                  sc["labor"],             "normal",   "sum",    "#,##0"),
        ("  その他販管費",                                        sc["other"],             "normal",   "sum",    "#,##0"),
        ("", None, "blank", None, None),

        # ── バズ→ENTIAL ──────────────────────────
        ("【バズ → ENTIAL 成功報酬】",                             None,                    "sec",      None,     "#,##0"),
        ("  ENTIAL業務委託費（TIE×5%）",                        c_data["ential_fee"],    "normal",   "sum",    "#,##0"),
        ("", None, "blank", None, None),

        # ── 販管費合計・手残り ────────────────────
        ("販売管理費合計",                                        c_data["sga"],           "total_cost","sum",   "#,##0"),
        ("", None, "blank", None, None),
        ("バズ 営業利益（手残り）",                               c_data["op_profit"],     "op",       "sum",    "#,##0"),
        ("営業利益率",                                            c_data["op_rate"],       "rate",     "avg_nz", None),
    ]

    for (label, vals, style, agg_method, num_fmt) in rows:
        ROW += 1
        ws.row_dimensions[ROW].height = (6 if style == "blank"
                                         else 15 if style == "kpi_sub" else 18)
        st = STYLE[style]
        cell = ws.cell(row=ROW, column=1, value=label)
        cell.font = Font(name="Meiryo UI", bold=st["bold"], size=st["sz"], color=st["fg"])
        cell.fill = PatternFill("solid", fgColor=st["bg"])
        cell.alignment = Alignment(horizontal=st["ha"], vertical="center", indent=1)

        if vals is None:
            fill_empty_row(ws, ROW, 2, 14, st["bg"])
            continue

        is_rate = (style == "rate")
        cell_fmt = ('0.0"%"' if is_rate else (num_fmt or "#,##0"))

        for j, v in enumerate(vals):
            col = 2 + j
            disp_val = "-" if (v == 0 and style == "kpi_sub") else v
            cell2 = ws.cell(row=ROW, column=col, value=disp_val)
            if style == "op" and j == bep:
                cell2.fill = PatternFill("solid", fgColor=GREEN)
                cell2.font = Font(name="Meiryo UI", bold=True, size=st["sz"], color=WHITE)
            else:
                cell2.fill = PatternFill("solid", fgColor=st["bg"])
                fg_v = (WHITE if style == "op" and v >= 0
                        else "FFCCCC" if style == "op" and v < 0
                        else "145A32" if (style in ("gross", "ec_gross", "tie_gross") and v >= 0)
                        else RED_ACC if (style in ("gross", "ec_gross", "tie_gross") and v < 0)
                        else st["fg"])
                cell2.font = Font(name="Meiryo UI", bold=st["bold"], size=st["sz"], color=fg_v)
            cell2.alignment = Alignment(horizontal="center", vertical="center")
            if disp_val != "-":
                cell2.number_format = cell_fmt

        non_z2 = [v for v in vals if v != 0]
        fmt2 = '0.0"%"' if is_rate else (num_fmt or "#,##0")
        if agg_method == "avg_nz":
            agg_v = round(sum(non_z2)/len(non_z2), 1) if non_z2 else 0.0
        else:
            agg_v = sum(vals)

        disp_agg = "-" if (agg_v == 0 and style == "kpi_sub") else agg_v
        cell3 = ws.cell(row=ROW, column=14, value=disp_agg)
        fg3 = (WHITE if style == "op" else
               ("145A32" if (style in ("gross", "ec_gross", "tie_gross") and agg_v >= 0)
                else RED_ACC if (style in ("gross", "ec_gross", "tie_gross") and agg_v < 0)
                else st["fg"]))
        bg3 = (GREEN if (style == "op" and agg_v >= 0)
               else RED_ACC if (style == "op" and agg_v < 0)
               else st["bg"])
        cell3.font = Font(name="Meiryo UI", bold=True, size=st["sz"], color=fg3)
        cell3.fill = PatternFill("solid", fgColor=bg3)
        cell3.alignment = Alignment(horizontal="center", vertical="center")
        if disp_agg != "-":
            cell3.number_format = fmt2

    thin = Side(style="thin", color="DDDDDD")
    for row in ws.iter_rows(min_row=HDR_ROW, max_row=ROW, min_col=1, max_col=14):
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws.freeze_panes = "B4"


# ═══════════════════════════════════════════════════════
# メイン実行
# ═══════════════════════════════════════════════════════
wb = openpyxl.Workbook()
default_ws = wb.active

# 比較シートを最初に
ws_cmp = wb.create_sheet(title="比較・撤退基準", index=0)
build_comparison_sheet(ws_cmp)

# 各シナリオのPLシートと内訳シートを生成
for sc in SCENARIOS:
    ws = wb.create_sheet(title=sc["sheet"])
    build_pl_sheet(ws, sc)
    ws2 = wb.create_sheet(title=sc["sheet"] + "_内訳")
    build_breakdown_sheet(ws2, sc)

# デフォルトシート削除
wb.remove(default_ws)

OUTPUT = "PL_3シナリオ比較.xlsx"
wb.save(OUTPUT)

print(f"\n✅ 完成: {OUTPUT}")
for sc in SCENARIOS:
    c = sc["c"]
    bep = f"M{c['bep_month']+1}" if c["bep_month"] is not None else "達成せず"
    print(f"  [{sc['sheet']}] 売上={sum(c['rev_total']):>7,}万  "
          f"営業利益={sum(c['op_profit']):>+7,}万  BEP={bep}")
