"""
Year1 月次PL (①大成功パターン) v2
- バズ財務データ参考反映
- KSF/KPI埋め込み
- ピンクなしコーポレートブルー系
- 単価仮置き明記
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Year1_月次PL_大成功"
ws.sheet_view.showGridLines = False

# ── カラー（コーポレートネイビー系）──────────────────────
NAVY       = "1E3A8A"   # ヘッダー
NAVY_DK    = "0F172A"   # 強調
BLUE       = "2563EB"   # 売上セクション
BROWN      = "92400E"   # 原価セクション
PURPLE     = "6B21A8"   # 販管費セクション
GREEN      = "065F46"   # 利益セクション
ROW_REV    = "EFF6FF"   # 売上行背景
ROW_COGS   = "FEF3C7"   # 原価行背景
ROW_SGA    = "F5F3FF"   # 販管費行背景
ROW_PROF   = "ECFDF5"   # 利益行背景
ROW_KPI    = "F1F5F9"   # KPI行背景
TOTAL_BG   = "D1FAE5"   # 合計行背景
WHITE      = "FFFFFF"
GRAY_TEXT  = "6B7280"
DARK_TEXT  = "111827"
LABEL_BG   = "F9FAFB"
BORDER     = "E5E7EB"
NEG_COLOR  = "DC2626"
POS_COLOR  = "059669"

def f(c): return PatternFill("solid", fgColor=c)
def b(color=BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def setc(r, c, val, bold=False, size=9, bg=None, align="left",
         fmt=None, italic=False, color=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(bold=bold, size=size,
                     color=color or DARK_TEXT,
                     italic=italic, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if bg: cell.fill = f(bg)
    if fmt: cell.number_format = fmt
    cell.border = b()
    return cell

def merge_fill(rng, val, bg, bold=True, size=11, color=WHITE, align="left", italic=False):
    ws.merge_cells(rng)
    start = rng.split(":")[0]
    col = ord(start[0].upper()) - ord("A") + 1 if start[0].isalpha() else 1
    row = int("".join(c for c in start if c.isdigit()))
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")
    cell.fill = f(bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")

# ── 列幅 ──────────────────────────────────────────────────
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 30
for i in range(3, 15):  # M1〜M12
    ws.column_dimensions[get_column_letter(i)].width = 10
ws.column_dimensions["O"].width = 12  # 通期
ws.column_dimensions["P"].width = 35  # 備考

# ══════════════════════════════════════════════════════════
# 数値設定（①大成功パターン）
# ══════════════════════════════════════════════════════════

M = 12

# ── 売上（単価仮置き）─────────────────────────────────────
# ① インバウンド体験：@20,000円/人（M1-3仕込み、M4から開始）
inbound_pax   = [0, 0, 0, 50, 80, 100, 150, 200, 230, 280, 320, 380]
inbound_price = 2.0  # 万円/人
inbound_rev   = [p * inbound_price for p in inbound_pax]

# ② 広告タイアップ：@300万/件
taieup_count  = [0, 0, 0, 1, 1, 2, 2, 3, 3, 4, 4, 5]
taieup_price  = 300  # 万円/件
taieup_rev    = [c * taieup_price for c in taieup_count]

# ③ WESELL グッズ売上（米国TikTok Shop）
wesell_rev    = [0, 0, 0, 0, 50, 100, 150, 200, 250, 350, 450, 550]

# ④ IPライセンス：@50万/社/月
license_cos   = [0, 0, 0, 0, 0, 1, 1, 2, 2, 3, 3, 4]  # 契約社数
license_price = 50  # 万円/社/月
license_rev   = [c * license_price for c in license_cos]

# ⑤ ギャル旅タイアップ：@400万/件 年2件
galtrip_rev   = [0, 0, 0, 0, 400, 0, 0, 0, 400, 0, 0, 0]

total_rev = [sum(x) for x in zip(
    inbound_rev, taieup_rev, wesell_rev, license_rev, galtrip_rev)]

# ── 売上原価（バズ財務データ参考）─────────────────────────
# 【バズ参考】SMM粗利率34.3% / ソーシャルベース粗利率51-57%
inbound_cogs = [r * 0.30 for r in inbound_rev]   # 体験30%（会場・衣装・スタッフ）
taieup_cogs  = [r * 0.40 for r in taieup_rev]    # タイアップ40%（制作・インフルエンサー費）
wesell_cogs  = [r * 0.45 for r in wesell_rev]    # WESELL45%（仕入原価）
license_cogs = [r * 0.05 for r in license_rev]   # ライセンス5%
galtrip_cogs = [r * 0.35 for r in galtrip_rev]   # ギャル旅35%

total_cogs   = [sum(x) for x in zip(
    inbound_cogs, taieup_cogs, wesell_cogs, license_cogs, galtrip_cogs)]
gross_profit = [r - c for r, c in zip(total_rev, total_cogs)]
gp_margin    = [gp/r if r > 0 else 0 for gp, r in zip(gross_profit, total_rev)]

# ── 販管費（バズ財務構成参考）──────────────────────────
# 【バズ参考】販管費内訳：給与34%、業務委託4%、支払手数料4%、広告宣伝費1.2%
content_cost = [50, 100, 150, 150, 200, 200, 250, 250, 300, 300, 300, 300]   # 業務委託ベース
gyaru_share  = [round((taieup_rev[i]*0.32 + license_rev[i]*0.50)) for i in range(M)]
ential_fixed = [80]*M  # バズ社内PM固定
ential_inc   = [round(r * 0.03) for r in total_rev]  # 売上連動
adv_cost     = [30, 30, 50, 50, 80, 80, 80, 100, 100, 100, 120, 120]
other_cost   = [30]*M  # 通信・交通・雑費

total_sga = [sum(x) for x in zip(
    content_cost, gyaru_share, ential_fixed, ential_inc, adv_cost, other_cost)]

# ── 営業利益 ─────────────────────────────────────────────
op_profit = [gp - sga for gp, sga in zip(gross_profit, total_sga)]
op_margin = [op/r if r > 0 else 0 for op, r in zip(op_profit, total_rev)]

cumulative = []
cum = 0
for v in op_profit:
    cum += v
    cumulative.append(cum)

# ══════════════════════════════════════════════════════════
# Excel出力
# ══════════════════════════════════════════════════════════

R = 1

# タイトル
merge_fill(f"A{R}:P{R}",
    "損益計算書（P/L）　Year1 月次　①大成功パターン　単位：万円",
    NAVY, size=13, align="center")
ws.row_dimensions[R].height = 28
R += 1

# サブタイトル
merge_fill(f"A{R}:P{R}",
    "【前提】TikTokフォロワー急伸、M4から受注開始　／　"
    "単価は仮置き（セル直接修正可）　／　サイバーバズ財務データ（SMM粗利率34.3%、SB粗利率51%）を参考",
    LABEL_BG, size=8, color=GRAY_TEXT, bold=False, align="left")
ws.row_dimensions[R].height = 18
R += 2

# ── 列ヘッダー ──────────────────────────────────────────
months = [f"M{i+1}" for i in range(M)]
header = ["", "項　目"] + months + ["通期合計", "備考"]
for j, h in enumerate(header):
    cell = ws.cell(row=R, column=j+1, value=h)
    cell.font = Font(bold=True, size=9, color=WHITE, name="Calibri")
    cell.fill = f(NAVY)
    cell.alignment = Alignment(
        horizontal="center" if j > 1 else "left",
        vertical="center")
    cell.border = b(NAVY)
ws.row_dimensions[R].height = 22
R += 1

# ── セクション見出し ────────────────────────────────────
def section(row, label, bg):
    ws.merge_cells(f"A{row}:P{row}")
    cell = ws.cell(row=row, column=1, value=f"  {label}")
    cell.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
    cell.fill = f(bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = b(bg)
    ws.row_dimensions[row].height = 20

# ── データ行 ────────────────────────────────────────────
def drow(row, indent, label, values, bg, note="",
         bold=False, is_pct=False, is_count=False):
    setc(row, 1, "", bg=bg)
    setc(row, 2, ("    " * indent) + label, bold=bold, bg=bg,
         color=DARK_TEXT if bold else "374151")
    annual = sum(values)
    for i, v in enumerate(values):
        col = 3 + i
        if is_pct:
            setc(row, col, f"{v*100:.1f}%" if v != 0 else "-",
                 bg=bg, align="right",
                 color=POS_COLOR if v > 0 else NEG_COLOR if v < 0 else GRAY_TEXT)
        elif is_count:
            setc(row, col, v if v > 0 else "-",
                 bg=bg, align="right", color=GRAY_TEXT)
        else:
            setc(row, col, round(v, 1) if v != 0 else "-",
                 bg=bg, align="right", fmt="#,##0",
                 color=NEG_COLOR if v < 0 else (POS_COLOR if bold and v > 0 else DARK_TEXT),
                 bold=bold)
    # 通期合計
    if is_pct:
        yr_pct = annual / M if annual > 0 else 0  # 年平均（参考値）
        setc(row, 15, "-", bg=bg, align="right", color=GRAY_TEXT, italic=True)
    elif is_count:
        setc(row, 15, annual, bg=bg, align="right", bold=bold, color=GRAY_TEXT)
    else:
        setc(row, 15, round(annual, 0),
             bg=bg, align="right", fmt="#,##0",
             color=NEG_COLOR if annual < 0 else DARK_TEXT,
             bold=bold)
    # 備考
    setc(row, 16, note, bg=bg, align="left",
         color=GRAY_TEXT, italic=True, size=8)
    ws.row_dimensions[row].height = 17

# ══════════════════════════════════════════════════════════
# ▼ 売上高
# ══════════════════════════════════════════════════════════
section(R, "▼ 売上高（Revenue）", BLUE); R += 1

drow(R, 1, "① インバウンド体験売上", inbound_rev, ROW_REV,
     note="渋谷ギャル体験（ppgalclub参考）"); R += 1
drow(R, 2, "　└ 月間体験人数（人）", inbound_pax, ROW_REV,
     note="KPI：月50→380人", is_count=True); R += 1
drow(R, 2, "　└ 単価（万円/人）", [inbound_price]*M if True else [], ROW_REV,
     note="仮置き：¥20,000/人", is_count=True); R += 1

drow(R, 1, "② 広告タイアップ売上", taieup_rev, ROW_REV,
     note="企業×TikTokショートドラマ"); R += 1
drow(R, 2, "　└ 月間案件数（件）", taieup_count, ROW_REV,
     note="KPI：月1→5件", is_count=True); R += 1
drow(R, 2, "　└ 単価（万円/件）", [taieup_price]*M, ROW_REV,
     note="仮置き：300万/件", is_count=True); R += 1

drow(R, 1, "③ WESELL グッズ売上", wesell_rev, ROW_REV,
     note="米国TikTok Shop / バズWESELL運用"); R += 1

drow(R, 1, "④ IPライセンス売上", license_rev, ROW_REV,
     note="ギャル協会公認ロゴ等"); R += 1
drow(R, 2, "　└ 契約社数", license_cos, ROW_REV,
     note="KPI：0→4社", is_count=True); R += 1
drow(R, 2, "　└ 単価（万円/社/月）", [license_price]*M, ROW_REV,
     note="仮置き：月50万", is_count=True); R += 1

drow(R, 1, "⑤ ギャル旅タイアップ", galtrip_rev, ROW_REV,
     note="自治体・観光協会、年2件、@400万"); R += 1

drow(R, 0, "  売上高　合計", total_rev, TOTAL_BG, bold=True,
     note="Year1通期目標：1.5億円〜"); R += 1
R += 1

# ══════════════════════════════════════════════════════════
# ▼ 売上原価
# ══════════════════════════════════════════════════════════
section(R, "▼ 売上原価（COGS）　※バズ財務データ参考で原価率設定", BROWN); R += 1

drow(R, 1, "インバウンド体験原価（会場・衣装・スタッフ）",
     inbound_cogs, ROW_COGS, note="売上の30%（仮置き）"); R += 1
drow(R, 1, "広告タイアップ原価（制作・インフルエンサー）",
     taieup_cogs, ROW_COGS, note="売上の40%（バズSMM参考）"); R += 1
drow(R, 1, "WESELL仕入原価",
     wesell_cogs, ROW_COGS, note="売上の45%（バズSB参考）"); R += 1
drow(R, 1, "IPライセンス原価",
     license_cogs, ROW_COGS, note="売上の5%（ほぼ固定費）"); R += 1
drow(R, 1, "ギャル旅制作費",
     galtrip_cogs, ROW_COGS, note="売上の35%"); R += 1

drow(R, 0, "  売上原価　合計", total_cogs, TOTAL_BG, bold=True); R += 1
drow(R, 0, "  売上総利益（Gross Profit）", gross_profit, TOTAL_BG,
     bold=True, note="売上−原価"); R += 1
drow(R, 1, "  売上総利益率（GP Margin %）", gp_margin, TOTAL_BG,
     is_pct=True, note="バズ連結参考：34.3%"); R += 1
R += 1

# ══════════════════════════════════════════════════════════
# ▼ 販売費及び一般管理費
# ══════════════════════════════════════════════════════════
section(R, "▼ 販売費及び一般管理費（SG&A）　※バズ販管費構成参考", PURPLE); R += 1

drow(R, 1, "コンテンツ制作費（TikTok/IG動画）",
     content_cost, ROW_SGA, note="業務委託費ベース（バズ参考）"); R += 1
drow(R, 1, "ギャル協会レベニューシェア",
     gyaru_share, ROW_SGA, note="タイアップ32%＋ライセンス50%"); R += 1
drow(R, 1, "ENTIAL固定費（バズ社内PM）",
     ential_fixed, ROW_SGA, note="月80万固定（バズの人件費ゼロ）"); R += 1
drow(R, 1, "ENTIALインセンティブ",
     ential_inc, ROW_SGA, note="売上の3%（成果連動）"); R += 1
drow(R, 1, "広告宣伝費",
     adv_cost, ROW_SGA, note="SNS広告・PR費（バズ販管費1.2%参考）"); R += 1
drow(R, 1, "その他運営費",
     other_cost, ROW_SGA, note="通信・交通・雑費"); R += 1

drow(R, 0, "  販管費　合計", total_sga, TOTAL_BG, bold=True); R += 1
R += 1

# ══════════════════════════════════════════════════════════
# ▼ 営業利益
# ══════════════════════════════════════════════════════════
section(R, "▼ 営業利益（Operating Profit）", GREEN); R += 1

drow(R, 0, "  営業利益", op_profit, ROW_PROF, bold=True,
     note="売上総利益−販管費"); R += 1
drow(R, 1, "  営業利益率（Operating Margin %）",
     op_margin, ROW_PROF, is_pct=True,
     note="バズ連結参考：6.7%"); R += 1
drow(R, 1, "  累計営業利益（BS）", cumulative, ROW_PROF,
     note="累計＝損益分岐月を示す"); R += 1
R += 1

# ══════════════════════════════════════════════════════════
# ▼ KSF／KPI（成功の鍵となる指標）
# ══════════════════════════════════════════════════════════
section(R, "▼ KSF／KPI（成功の鍵）", NAVY_DK); R += 1

drow(R, 1, "TikTokフォロワー数（千人）",
     [0, 1, 3, 5, 8, 15, 25, 40, 60, 80, 100, 130],
     ROW_KPI, note="Year1末：13万人目標", is_count=True); R += 1
drow(R, 1, "月間動画再生回数（百万回）",
     [0, 0.5, 1, 3, 5, 10, 15, 25, 40, 60, 80, 100],
     ROW_KPI, note="バズり指標", is_count=True); R += 1
drow(R, 1, "インバウンド体験 月間人数",
     inbound_pax, ROW_KPI,
     note="KSF①：リピート×口コミ", is_count=True); R += 1
drow(R, 1, "インバウンド体験 客単価（千円）",
     [0, 0, 0, 20, 20, 20, 20, 22, 22, 25, 25, 25],
     ROW_KPI, note="オプション追加で向上", is_count=True); R += 1
drow(R, 1, "月間タイアップ案件数",
     taieup_count, ROW_KPI,
     note="KSF②：ENTIAL営業力", is_count=True); R += 1
drow(R, 1, "タイアップ平均単価（万円/件）",
     [0, 0, 0, 300, 300, 300, 350, 350, 350, 400, 400, 400],
     ROW_KPI, note="実績で単価上昇", is_count=True); R += 1
drow(R, 1, "WESELL月間注文数（件）",
     [0, 0, 0, 0, 100, 200, 300, 400, 500, 700, 900, 1100],
     ROW_KPI, note="KSF③：米国TikTok", is_count=True); R += 1
drow(R, 1, "IPライセンス契約社数（累計）",
     license_cos, ROW_KPI,
     note="KSF④：IP価値向上", is_count=True); R += 1
R += 1

# ══════════════════════════════════════════════════════════
# ▼ 通期サマリー
# ══════════════════════════════════════════════════════════
section(R, "▼ Year1通期サマリー", NAVY_DK); R += 1

summary = [
    ("売上合計", f"{sum(total_rev):,.0f} 万円",
     f"= {sum(total_rev)/10000:.2f}億円"),
    ("売上総利益", f"{sum(gross_profit):,.0f} 万円",
     f"GP率 {sum(gross_profit)/sum(total_rev)*100:.1f}%"),
    ("販管費合計", f"{sum(total_sga):,.0f} 万円", ""),
    ("営業利益", f"{sum(op_profit):,.0f} 万円",
     f"OPM {sum(op_profit)/sum(total_rev)*100:.1f}%"),
    ("損益分岐月",
     f"M{next((i+1 for i,v in enumerate(cumulative) if v>=0), '-')}", ""),
]

for i, (label, val, note) in enumerate(summary):
    setc(R, 2, label, bold=True, size=10, bg=LABEL_BG)
    setc(R, 3, val, bold=True, size=11, bg=LABEL_BG, align="right",
         color=POS_COLOR if "利益" in label or "売上" in label else DARK_TEXT)
    ws.merge_cells(start_row=R, start_column=3, end_row=R, end_column=7)
    setc(R, 8, note, size=9, bg=LABEL_BG, color=GRAY_TEXT, italic=True)
    ws.merge_cells(start_row=R, start_column=8, end_row=R, end_column=16)
    ws.row_dimensions[R].height = 22
    R += 1

R += 1
# 注記
merge_fill(f"A{R}:P{R}",
    "【注記】①単価は仮置き　②原価率・販管費構成はバズFY26着予データ参考　"
    "③ENTIALはバズ社内PMのため追加人件費ゼロ　④数値はセル直接編集可",
    "F3F4F6", size=8, color=GRAY_TEXT, bold=False, italic=True, align="left")
ws.row_dimensions[R].height = 18

# ── 保存 ──────────────────────────────────────────────────
out = "/home/user/gyal-kyokai--/PL_Year1月次_①大成功.xlsx"
wb.save(out)

# コンソール出力
print(f"Saved: {out}")
print(f"\n【Year1 通期サマリー（①大成功）】")
print(f"  売上合計：    {sum(total_rev):>8,.0f} 万円 ({sum(total_rev)/10000:.2f}億円)")
print(f"  売上総利益：  {sum(gross_profit):>8,.0f} 万円 (GP率 {sum(gross_profit)/sum(total_rev)*100:.1f}%)")
print(f"  販管費合計：  {sum(total_sga):>8,.0f} 万円")
print(f"  営業利益：    {sum(op_profit):>8,.0f} 万円 (OPM {sum(op_profit)/sum(total_rev)*100:.1f}%)")
print(f"  損益分岐月：  M{next((i+1 for i,v in enumerate(cumulative) if v>=0), '-')}")
