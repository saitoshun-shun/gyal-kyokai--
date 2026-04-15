"""
事業計画書 - バズ社PL Year1月次（❶大成功パターン）
ギャル×インバウンド共同事業
  軸①: Instagram運用 × 体験ツアータイアップ
  軸②: TikTok Shop（WESELL）× ギャルグッズEC
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Year1月次PL_大成功"

# ─── カラー定義（シンプル） ───
NAVY    = "1A1A2E"
NAVY2   = "16213E"
BLUE    = "2E4A7A"
RED_ACC = "C0392B"
GREEN   = "1E8449"
LT_BLUE = "D6EAF8"
LT_GRN  = "D5F5E3"
LT_RED  = "FADBD8"
LT_GRY  = "F2F3F4"
MID_GRY = "E5E7E9"
WHITE   = "FFFFFF"
GRY_TXT = "7F8C8D"

# ─── ヘルパー ───
def c(row, col, val="", bold=False, sz=10, fg="222222", bg=None,
      ha="center", wrap=False, italic=False, fmt=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(name="Meiryo UI", bold=bold, size=sz,
                     color=fg, italic=italic)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=ha, vertical="center",
                               wrap_text=wrap)
    if fmt:
        cell.number_format = fmt
    return cell

def border_all(row, col, color="CCCCCC"):
    s = Side(style="thin", color=color)
    ws.cell(row=row, column=col).border = Border(
        top=s, bottom=s, left=s, right=s)

# ─── 列・行レイアウト ───
# 列: A=項目, B〜M=M1〜M12, N=Year1合計
NCOL_ITEM  = 1
NCOL_M1    = 2
NCOL_TOTAL = 14
MONTHS = [f"M{i}" for i in range(1, 13)]

ws.column_dimensions["A"].width = 32
for col in range(2, 15):
    ws.column_dimensions[get_column_letter(col)].width = 10

# ─── 数値定義（大成功パターン） ───────────────────────────
# ■ 収益
# IG軸：体験タイアップ＆メディアフィー（バズへの受注額 / 万円）
# M8で損益分岐、Year1合計≒1.8億を目指す設計
ig_tie   = [0, 0, 150, 250, 400, 600, 800, 1000, 1200, 1500, 1800, 2200]

# TikTok軸：グッズEC グロス売上（バズがWESELL窓口として計上 / 万円）
# 件数 × 平均単価 で算出
ec_units = [0, 0, 50, 120, 250, 450, 700, 1000, 1300, 1700, 2200, 2800]  # 注文件数
ec_atp   = [0, 0, 5000, 5500, 6000, 6500, 7000, 7200, 7500, 7800, 8000, 8200]  # 平均単価(円)
ec_gross = [round(u * a / 10000) for u, a in zip(ec_units, ec_atp)]
# [0, 0, 25, 66, 150, 293, 490, 720, 975, 1326, 1760, 2296]

rev_total = [ig + ec for ig, ec in zip(ig_tie, ec_gross)]

# ■ 原価
ig_cogs  = [round(v * 0.35) for v in ig_tie]          # コンテンツ制作費 35%
ec_cogs  = [round(v * 0.45) for v in ec_gross]         # EC商品原価 45%
gal_rs   = [round(ig * 0.30 + ec * 0.15)               # ギャル協会レベニューシェア
            for ig, ec in zip(ig_tie, ec_gross)]        # IG×30% + EC×15%
cogs_total = [ic + eco + g for ic, eco, g in zip(ig_cogs, ec_cogs, gal_rs)]
gross    = [r - cg for r, cg in zip(rev_total, cogs_total)]

# ■ 販売管理費
ad_exp   = [100,150,200,200,250,250,300,300,300,300,300,300]  # 広告費
labor    = [100,100,100,100,100,100,100,100,100,100,100,100]  # 人件費（バズ社員）
wesell   = [0,  0,  5,  10, 20, 35, 50, 70, 90,110,130,150]  # WESELL利用料等
other    = [50, 50, 50, 50, 60, 60, 70, 70, 80, 80, 90, 90]  # その他
sga_total= [a+l+w+o for a,l,w,o in zip(ad_exp,labor,wesell,other)]

# ■ 営業利益
op_profit= [g - s for g, s in zip(gross, sga_total)]

# 損益分岐月を取得
bep_month = next((i for i, v in enumerate(op_profit) if v > 0), None)

# ─────────────────────────────────────────────────────
# SECTION 0: タイトル
# ─────────────────────────────────────────────────────
ROW = 1
ws.row_dimensions[ROW].height = 38
ws.merge_cells(f"A{ROW}:N{ROW}")
cell = ws.cell(row=ROW, column=1,
               value="❶ 大成功パターン  ／  Year1 月次事業計画（バズ社PL）")
cell.font = Font(name="Meiryo UI", bold=True, size=16, color=WHITE)
cell.fill = PatternFill("solid", fgColor=NAVY)
cell.alignment = Alignment(horizontal="center", vertical="center")

ROW += 1
ws.row_dimensions[ROW].height = 16
ws.merge_cells(f"A{ROW}:N{ROW}")
cell = ws.cell(row=ROW, column=1,
               value="ギャル×インバウンド共同事業（IG体験タイアップ × TikTok Shop WESELL）　2026年4月〜2027年3月　単位：万円")
cell.font = Font(name="Meiryo UI", size=9, color="DDDDDD")
cell.fill = PatternFill("solid", fgColor=NAVY2)
cell.alignment = Alignment(horizontal="center", vertical="center")

# ─────────────────────────────────────────────────────
# SECTION 1: ヘッダー行
# ─────────────────────────────────────────────────────
ROW += 1
ws.row_dimensions[ROW].height = 26
c(ROW, NCOL_ITEM, "項目", bold=True, sz=10, fg=WHITE, bg=BLUE)
for i, m in enumerate(MONTHS):
    c(ROW, NCOL_M1+i, m, bold=True, sz=10, fg=WHITE, bg=BLUE)
c(ROW, NCOL_TOTAL, "Year1合計", bold=True, sz=10, fg=WHITE, bg=BLUE)
HDR_ROW = ROW

# ─────────────────────────────────────────────────────
# SECTION 2: KPIセクション
# ─────────────────────────────────────────────────────
ROW += 1
ws.row_dimensions[ROW].height = 22
ws.merge_cells(f"A{ROW}:N{ROW}")
cell = ws.cell(row=ROW, column=1, value="▌ 重要KPI（月次推移）")
cell.font = Font(name="Meiryo UI", bold=True, size=11, color=WHITE)
cell.fill = PatternFill("solid", fgColor=BLUE)
cell.alignment = Alignment(horizontal="left", vertical="center",
                           indent=1)

kpi_rows = [
    # (ラベル, データリスト, 書式, ★KPIマーク, 合計方法)
    # 合計方法: "last"=最終値, "sum"=合計, "avg"=平均
    ("IGフォロワー数（累計）",
     [500,2000,5000,10000,18000,28000,40000,55000,72000,92000,115000,150000],
     "#,##0", True, "last"),
    ("TikTokフォロワー数（累計）",
     [300,1200,3000,7000,14000,23000,35000,50000,68000,88000,110000,140000],
     "#,##0", True, "last"),
    ("IG投稿リーチ数（月間・万）",
     [1,3,8,15,25,40,55,70,85,100,120,150],
     "#,##0", False, "avg"),
    ("体験ツアー予約件数（月次）",
     [0,0,2,4,7,10,13,16,20,24,28,32],
     "#,##0", True, "sum"),
    ("体験ツアー平均単価（万円）",
     [0,0,15,17,18,20,22,23,25,25,28,30],
     "#,##0", False, "avg"),
    ("EC注文件数（TikTok Shop）",
     ec_units,
     "#,##0", True, "sum"),
    ("EC平均注文単価（円）",
     ec_atp,
     "#,##0", False, "avg"),
    ("タイアップ受注件数（月次）",
     [0,0,1,1,2,3,3,4,4,5,6,7],
     "#,##0", False, "sum"),
]

for ki, (label, vals, fmt, is_kpi, agg) in enumerate(kpi_rows):
    ROW += 1
    ws.row_dimensions[ROW].height = 17
    row_bg = LT_GRY if ki % 2 == 0 else WHITE
    # 項目列
    cell = ws.cell(row=ROW, column=NCOL_ITEM, value=label)
    cell.font = Font(name="Meiryo UI", size=9, bold=is_kpi,
                     color=RED_ACC if is_kpi else "333333")
    cell.fill = PatternFill("solid", fgColor=row_bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # データ列
    for j, v in enumerate(vals):
        cell = ws.cell(row=ROW, column=NCOL_M1+j, value=v if v != 0 else "-")
        cell.font = Font(name="Meiryo UI", size=9, bold=is_kpi,
                         color=RED_ACC if is_kpi else "444444")
        cell.fill = PatternFill("solid", fgColor=row_bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if v != 0:
            cell.number_format = fmt

    # 合計列
    if agg == "last":
        agg_val = vals[-1]
    elif agg == "sum":
        agg_val = sum(vals)
    else:
        non_zero = [v for v in vals if v != 0]
        agg_val = round(sum(non_zero)/len(non_zero)) if non_zero else 0

    cell = ws.cell(row=ROW, column=NCOL_TOTAL, value=agg_val)
    cell.font = Font(name="Meiryo UI", size=9, bold=True,
                     color=RED_ACC if is_kpi else "333333")
    cell.fill = PatternFill("solid", fgColor=row_bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt

KPI_END_ROW = ROW

# ─────────────────────────────────────────────────────
# SECTION 3: PLセクション ヘッダー
# ─────────────────────────────────────────────────────
ROW += 1
ws.row_dimensions[ROW].height = 22
ws.merge_cells(f"A{ROW}:N{ROW}")
cell = ws.cell(row=ROW, column=1, value="▌ 損益計算書（P/L）　単位：万円")
cell.font = Font(name="Meiryo UI", bold=True, size=11, color=WHITE)
cell.fill = PatternFill("solid", fgColor=NAVY)
cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# ─────────────────────────────────────────────────────
# PL行データ定義
# ─────────────────────────────────────────────────────
# style一覧:
#   "sec"      : セクション見出し（灰背景）
#   "normal"   : 通常行（白背景、インデント）
#   "subtotal" : 小計（ライトブルー）
#   "gross"    : 粗利（ライトグリーン）
#   "rate"     : 率（薄グレー、%表示）
#   "op"       : 営業利益（緑 or 赤）
#   "blank"    : 空行

gp_rate = [round(g/r*100, 1) if r > 0 else 0.0
           for g, r in zip(gross, rev_total)]
op_rate = [round(o/r*100, 1) if r > 0 else 0.0
           for o, r in zip(op_profit, rev_total)]

pl_rows = [
    ("【売上高】", None, "sec"),
    ("  ① IG体験タイアップ・メディアフィー", ig_tie, "normal"),
    ("  ② TikTok Shop / WESELL グッズEC",    ec_gross, "normal"),
    ("売上合計",  rev_total, "subtotal"),
    ("", None, "blank"),
    ("【売上原価】", None, "sec"),
    ("  コンテンツ制作費（IG軸 / 売上×35%）",  ig_cogs,    "normal"),
    ("  EC商品原価（TikTok軸 / EC売上×45%）",  ec_cogs,    "normal"),
    ("  ギャル協会レベニューシェア（IG×30%＋EC×15%）", gal_rs, "normal"),
    ("売上原価合計",  cogs_total, "subtotal"),
    ("粗利益",        gross,      "gross"),
    ("粗利率",        gp_rate,    "rate"),
    ("", None, "blank"),
    ("【販売管理費】", None, "sec"),
    ("  広告費（IG/TikTok広告）",          ad_exp,  "normal"),
    ("  人件費（バズ社員 / ENTIALはゼロ）", labor,   "normal"),
    ("  WESELL利用料・決済手数料",          wesell,  "normal"),
    ("  その他販管費",                      other,   "normal"),
    ("販売管理費合計", sga_total, "subtotal"),
    ("", None, "blank"),
    ("営業利益",    op_profit, "op"),
    ("営業利益率",  op_rate,   "rate"),
]

# styleごとの書式設定
STYLE = {
    "sec":      dict(bold=True,  sz=9,  fg="1A1A2E", bg=MID_GRY, ha="left"),
    "normal":   dict(bold=False, sz=9,  fg="333333", bg=WHITE,    ha="left"),
    "subtotal": dict(bold=True,  sz=9,  fg="1A3A6B", bg=LT_BLUE,  ha="left"),
    "gross":    dict(bold=True,  sz=10, fg="145A32", bg=LT_GRN,   ha="left"),
    "rate":     dict(bold=False, sz=9,  fg=GRY_TXT,  bg=LT_GRY,  ha="left"),
    "op":       dict(bold=True,  sz=11, fg="FFFFFF",  bg=NAVY,    ha="left"),
    "blank":    dict(bold=False, sz=8,  fg=WHITE,     bg=WHITE,   ha="left"),
}

for pi, (label, vals, style) in enumerate(pl_rows):
    ROW += 1
    ws.row_dimensions[ROW].height = 8 if style == "blank" else 18
    st = STYLE[style]

    # 項目列
    cell = ws.cell(row=ROW, column=NCOL_ITEM, value=label)
    cell.font = Font(name="Meiryo UI", bold=st["bold"], size=st["sz"],
                     color=st["fg"])
    cell.fill = PatternFill("solid", fgColor=st["bg"])
    cell.alignment = Alignment(horizontal=st["ha"], vertical="center",
                               indent=1)

    if vals is None:
        for col in range(NCOL_M1, NCOL_TOTAL+1):
            cell2 = ws.cell(row=ROW, column=col)
            cell2.fill = PatternFill("solid", fgColor=st["bg"])
        continue

    for j, v in enumerate(vals):
        col = NCOL_M1 + j
        cell2 = ws.cell(row=ROW, column=col, value=v)

        # 営業利益行：BEP月を強調
        if style == "op" and j == bep_month:
            cell2.fill = PatternFill("solid", fgColor=GREEN)
        else:
            cell2.fill = PatternFill("solid", fgColor=st["bg"])

        # 色分け
        if style == "op":
            fg_v = WHITE if v >= 0 else "FFCCCC"
        elif style in ("gross",):
            fg_v = "145A32" if v >= 0 else RED_ACC
        else:
            fg_v = st["fg"]

        cell2.font = Font(name="Meiryo UI", bold=st["bold"], size=st["sz"],
                          color=fg_v)
        cell2.alignment = Alignment(horizontal="center", vertical="center")
        if style == "rate":
            cell2.number_format = '0.0"%"'
        else:
            cell2.number_format = "#,##0"

    # 合計列
    col = NCOL_TOTAL
    if style == "rate":
        non_z = [v for v in vals if v != 0]
        agg_v = round(sum(non_z)/len(non_z), 1) if non_z else 0.0
        fmt_s = '0.0"%"'
    else:
        agg_v = sum(vals)
        fmt_s = "#,##0"

    cell3 = ws.cell(row=ROW, column=col, value=agg_v)
    cell3.font = Font(name="Meiryo UI", bold=True, size=st["sz"],
                      color=WHITE if style == "op" else st["fg"])
    cell3.fill = PatternFill("solid",
                             fgColor=GREEN if (style == "op" and agg_v >= 0)
                             else RED_ACC if (style == "op" and agg_v < 0)
                             else st["bg"])
    cell3.alignment = Alignment(horizontal="center", vertical="center")
    cell3.number_format = fmt_s

# ─────────────────────────────────────────────────────
# SECTION 4: BEP注記
# ─────────────────────────────────────────────────────
ROW += 1
ws.row_dimensions[ROW].height = 18
ws.merge_cells(f"A{ROW}:N{ROW}")
bep_label = f"M{bep_month+1}" if bep_month is not None else "Year1内なし"
cell = ws.cell(row=ROW, column=1,
               value=f"★ 損益分岐点：{bep_label}　から営業黒字転換（緑色セル）"
                     f"　／　Year1通期営業利益：{sum(op_profit):,}万円")
cell.font = Font(name="Meiryo UI", size=9, bold=True, color=NAVY)
cell.fill = PatternFill("solid", fgColor=LT_GRN)
cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# ─────────────────────────────────────────────────────
# SECTION 5: 注記フッター
# ─────────────────────────────────────────────────────
ROW += 1
ws.row_dimensions[ROW].height = 14
ws.merge_cells(f"A{ROW}:N{ROW}")
cell = ws.cell(row=ROW, column=1,
               value="【注記】 単位：万円　"
                     "／ IG原価率35%、EC原価率45%　"
                     "／ ギャル協会RS：IG売上×30%＋EC売上×15%　"
                     "／ ENTIALへの費用はバズPLに計上なし（別途協議）　"
                     "／ 人件費はバズ社員分のみ")
cell.font = Font(name="Meiryo UI", size=8, color=GRY_TXT, italic=True)
cell.fill = PatternFill("solid", fgColor=LT_GRY)
cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# ─────────────────────────────────────────────────────
# 枠線（KPI・PLの全セルに薄いborder）
# ─────────────────────────────────────────────────────
thin = Side(style="thin", color="DDDDDD")
for row in ws.iter_rows(min_row=HDR_ROW, max_row=ROW,
                        min_col=1, max_col=NCOL_TOTAL):
    for cell in row:
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

# ─────────────────────────────────────────────────────
# フリーズ・印刷設定
# ─────────────────────────────────────────────────────
ws.freeze_panes = "B4"

# 出力
OUTPUT = "PL_Year1月次_大成功_IG_TikTokShop.xlsx"
wb.save(OUTPUT)
print(f"✅ 完成: {OUTPUT}")
print(f"   Year1 売上合計    : {sum(rev_total):,} 万円")
print(f"   Year1 粗利        : {sum(gross):,} 万円（粗利率 {round(sum(gross)/sum(rev_total)*100,1)}%）")
print(f"   Year1 営業利益    : {sum(op_profit):,} 万円")
print(f"   損益分岐点        : M{bep_month+1}" if bep_month is not None else "   損益分岐点: Year1内になし")
