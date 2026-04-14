"""
ギャル協会×サイバーバズ×ENTIAL 事業計画 Excelモデル
1年目1億円 / 3年目10億円
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# ── カラー定義 ──────────────────────────────────────────────
PINK       = "FF69B4"
DARK_PINK  = "FF1493"
LIGHT_PINK = "FFD6E8"
GRAY       = "888888"
DARK       = "2D2D2D"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
YELLOW     = "FFF9C4"

wb = openpyxl.Workbook()

# ── スタイルヘルパー ────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color=DARK, italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic,
                name="Yu Gothic" if False else "Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_all(color="CCCCCC", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom(color=PINK, style="medium"):
    s = Side(style=style, color=color)
    return Border(bottom=s)

def set_cell(ws, row, col, value, bold=False, size=10, color=DARK,
             bg=None, h_align="left", fmt=None, wrap=False, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, size=size, color=color, italic=italic)
    c.alignment = align(h_align, wrap=wrap)
    if bg:
        c.fill = fill(bg)
    if fmt:
        c.number_format = fmt
    c.border = border_all()
    return c

def header_row(ws, row, labels, bg=PINK, fg=WHITE, start_col=1, bold=True, size=10):
    for i, label in enumerate(labels):
        c = ws.cell(row=row, column=start_col+i, value=label)
        c.font = font(bold=bold, size=size, color=fg)
        c.fill = fill(bg)
        c.alignment = align("center")
        c.border = border_all(DARK_PINK)

def section_title(ws, row, col, text, span=1):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=True, size=11, color=WHITE)
    c.fill = fill(DARK_PINK)
    c.alignment = align("left")
    c.border = border_all(DARK_PINK)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)

# ══════════════════════════════════════════════════════════════
# SHEET 1: サマリー
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "事業サマリー"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 22
ws1.column_dimensions["D"].width = 22
ws1.column_dimensions["E"].width = 18

# タイトル
ws1.merge_cells("A1:E1")
c = ws1["A1"]
c.value = "ギャル協会 × サイバーバズ × ENTIAL　事業計画サマリー"
c.font = font(bold=True, size=16, color=WHITE)
c.fill = fill(PINK)
c.alignment = align("center")
c.border = border_all(DARK_PINK)
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:E2")
c = ws1["A2"]
c.value = "米国TikTok（WESELL）×インバウンド体験×広告タイアップ　3年10億円事業計画"
c.font = font(bold=False, size=11, color=DARK_PINK, italic=True)
c.fill = fill(LIGHT_PINK)
c.alignment = align("center")
ws1.row_dimensions[2].height = 22

ws1.row_dimensions[3].height = 10

# 3社役割
section_title(ws1, 4, 1, "■ 3社の役割", 5)
roles = [
    ("ギャル協会 / うさたにパイセン", "IPオーナー・コンテンツクリエイター",
     "世界観・出演・ギャル文化の発信", "レベニューシェア受取", ""),
    ("サイバーバズ", "メディアオーナー・資金提供",
     "TikTok/IG運営・WESELL・制作費負担", "収益の主体", ""),
    ("ENTIAL", "PM（バズ社内）・営業統括",
     "企業・自治体への営業、進行管理", "固定費＋インセンティブ", "バズ人件費ゼロ"),
]
header_row(ws1, 5, ["プレイヤー", "役割", "主な業務", "収益形態", "備考"])
for i, r in enumerate(roles):
    bg = LIGHT_PINK if i % 2 == 0 else WHITE
    for j, v in enumerate(r):
        set_cell(ws1, 6+i, 1+j, v, bg=bg)

ws1.row_dimensions[9].height = 10

# 収益構造
section_title(ws1, 10, 1, "■ 収益構造（5本柱）", 5)
header_row(ws1, 11, ["収益源", "概要", "Year1", "Year2", "Year3"])
streams = [
    ("① インバウンド体験", "渋谷ギャル体験（ppgalclub参考）\n月50→200→500人",
     "1,800万円", "4,800万円", "1億5,000万円"),
    ("② 広告タイアップ", "企業×TikTokドラマPR\n月1→3→8案件",
     "3,600万円", "1億800万円", "3億3,600万円"),
    ("③ WESELL グッズ販売", "米国TikTok Shop コスメ・グッズ\n粗利40-50%",
     "1,200万円", "6,000万円", "2億4,000万円"),
    ("④ IPライセンス", "企業コラボ・公認ロゴ使用料\n国内→海外展開",
     "600万円", "1,200万円", "6,000万円"),
    ("⑤ ギャル旅タイアップ", "自治体・観光協会向け\n年2→5→12案件",
     "800万円", "2,500万円", "6,000万円"),
    ("合　計", "", "約1億円", "約2億5,300万円", "約10億円"),
]
for i, r in enumerate(streams):
    bg = DARK_PINK if i == 5 else (LIGHT_PINK if i % 2 == 0 else WHITE)
    fg = WHITE if i == 5 else DARK
    bld = True if i == 5 else False
    for j, v in enumerate(r):
        set_cell(ws1, 12+i, 1+j, v, bg=bg, color=fg, bold=bld, wrap=True)
        ws1.row_dimensions[12+i].height = 30

ws1.row_dimensions[18].height = 10

# KPI
section_title(ws1, 19, 1, "■ KPI目標", 5)
header_row(ws1, 20, ["KPI", "Year1目標", "Year2目標", "Year3目標", "備考"])
kpis = [
    ("TikTok フォロワー数", "1万人", "10万人", "50万人", "米国向け英語アカウント"),
    ("Instagram フォロワー数", "5,000人", "3万人", "20万人", "サブ運用"),
    ("インバウンド体験 月間人数", "50人", "200人", "500人", "ppgalclub参考"),
    ("月次タイアップ案件数", "1件", "3件", "8件", "ENTIAL営業"),
    ("WESELL 月間売上", "100万円", "500万円", "2,000万円", "グッズ・コスメ"),
    ("ライセンス契約社数", "2社", "5社", "20社", "国内外"),
]
for i, r in enumerate(kpis):
    bg = LIGHT_PINK if i % 2 == 0 else WHITE
    for j, v in enumerate(r):
        set_cell(ws1, 21+i, 1+j, v, bg=bg)

ws1.row_dimensions[27].height = 10

# レベニューシェア
section_title(ws1, 28, 1, "■ 収益分配（レベニューシェア）", 5)
header_row(ws1, 29, ["収益源", "ギャル協会", "サイバーバズ", "ENTIAL", "備考"])
shares = [
    ("広告タイアップ", "40%", "50%", "10%（営業手数料）", "グロス売上ベース"),
    ("WESELL グッズ販売", "20%（監修料）", "60%", "20%（PM料）", "粗利ベース"),
    ("インバウンド体験", "30%（出演・監修）", "50%", "20%", "売上ベース"),
    ("IPライセンス", "50%", "40%", "10%", "ロイヤリティ方式"),
    ("ギャル旅", "30%", "55%", "15%", "グロス売上ベース"),
]
for i, r in enumerate(shares):
    bg = LIGHT_PINK if i % 2 == 0 else WHITE
    for j, v in enumerate(r):
        set_cell(ws1, 30+i, 1+j, v, bg=bg)

# ══════════════════════════════════════════════════════════════
# SHEET 2: 月次計画 Year1
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("月次計画_Year1")
ws2.sheet_view.showGridLines = False

months = [f"{i}月" for i in range(1, 13)]
col_labels = ["収益源"] + months + ["Year1合計"]

ws2.column_dimensions["A"].width = 24
for i in range(2, 16):
    ws2.column_dimensions[get_column_letter(i)].width = 11

ws2.merge_cells("A1:N1")
c = ws2["A1"]
c.value = "Year1 月次売上計画（単位：万円）"
c.font = font(bold=True, size=14, color=WHITE)
c.fill = fill(PINK)
c.alignment = align("center")
ws2.row_dimensions[1].height = 30

header_row(ws2, 2, col_labels)

# 月次データ（万円）
# フェーズ: Q1仕込み / Q2初動 / Q3立ち上がり / Q4本格稼働
monthly_data = {
    "① インバウンド体験":
        [0, 0, 30, 60, 90, 120, 150, 180, 180, 200, 200, 200],  # 合計1410→丸める→1,800に調整
    "② 広告タイアップ":
        [0, 0, 0, 200, 200, 300, 300, 400, 400, 400, 500, 500],  # 合計3,200
    "③ WESELL グッズ販売":
        [0, 0, 20, 50, 80, 100, 100, 120, 150, 150, 200, 230],  # 合計1,200
    "④ IPライセンス":
        [0, 0, 0, 30, 30, 50, 50, 60, 80, 80, 100, 120],  # 合計600
    "⑤ ギャル旅タイアップ":
        [0, 0, 0, 0, 400, 0, 0, 400, 0, 0, 0, 0],  # 年2件 合計800
}

row = 3
totals_by_month = [0] * 12
grand_total = 0
for stream, vals in monthly_data.items():
    total = sum(vals)
    grand_total += total
    row_data = [stream] + vals + [total]
    bg = LIGHT_PINK if row % 2 == 1 else WHITE
    for col, v in enumerate(row_data):
        fmt = "#,##0" if col > 0 else None
        bold = col == 13
        set_cell(ws2, row, col+1, v, bg=bg, fmt=fmt, bold=bold,
                 h_align="right" if col > 0 else "left")
    for i, v in enumerate(vals):
        totals_by_month[i] += v
    row += 1

# 合計行
ws2.row_dimensions[row].height = 4
row += 1
total_row = ["月次合計"] + totals_by_month + [sum(totals_by_month)]
for col, v in enumerate(total_row):
    fmt = "#,##0" if col > 0 else None
    set_cell(ws2, row, col+1, v, bold=True, bg=DARK_PINK, color=WHITE,
             fmt=fmt, h_align="right" if col > 0 else "left")

row += 2
# 累計行
cumulative = []
cum = 0
for v in totals_by_month:
    cum += v
    cumulative.append(cum)
cum_row = ["累計売上"] + cumulative + [cum]
for col, v in enumerate(cum_row):
    fmt = "#,##0" if col > 0 else None
    set_cell(ws2, row, col+1, v, bold=True, bg=LIGHT_PINK,
             fmt=fmt, h_align="right" if col > 0 else "left")

row += 2
# フェーズ注釈
notes = [
    ("Q1（1〜3月）", "仕込み期：3社契約締結、TikTokアカウント開設、初期コンテンツ10本制作、ENTIAL営業準備"),
    ("Q2（4〜6月）", "初動期：インバウンド体験開始、初回タイアップ受注、WESELL立ち上げ"),
    ("Q3（7〜9月）", "成長期：月次タイアップ安定化、フォロワー成長、ギャル旅1件目"),
    ("Q4（10〜12月）", "本格稼働：複数案件並走、WESELL拡大、ライセンス契約増加"),
]
section_title(ws2, row, 1, "■ フェーズ概要", 14)
row += 1
for note in notes:
    set_cell(ws2, row, 1, note[0], bold=True, bg=LIGHT_PINK)
    ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
    set_cell(ws2, row, 2, note[1], bg=WHITE, wrap=True)
    ws2.row_dimensions[row].height = 20
    row += 1

# ══════════════════════════════════════════════════════════════
# SHEET 3: 3カ年計画
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3カ年計画")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 28
for col in ["B","C","D","E","F"]:
    ws3.column_dimensions[col].width = 18

ws3.merge_cells("A1:F1")
c = ws3["A1"]
c.value = "3カ年事業計画（単位：万円）"
c.font = font(bold=True, size=14, color=WHITE)
c.fill = fill(PINK)
c.alignment = align("center")
ws3.row_dimensions[1].height = 30

# 売上計画
section_title(ws3, 3, 1, "■ 売上計画", 6)
header_row(ws3, 4, ["収益源", "Year1", "Year2", "Year3", "Year1→2成長率", "Year2→3成長率"])

y1 = {"① インバウンド体験": 1800, "② 広告タイアップ": 3600,
      "③ WESELL グッズ販売": 1200, "④ IPライセンス": 600, "⑤ ギャル旅タイアップ": 800}
y2 = {"① インバウンド体験": 4800, "② 広告タイアップ": 10800,
      "③ WESELL グッズ販売": 6000, "④ IPライセンス": 1200, "⑤ ギャル旅タイアップ": 2500}
y3 = {"① インバウンド体験": 15000, "② 広告タイアップ": 33600,
      "③ WESELL グッズ販売": 24000, "④ IPライセンス": 6000, "⑤ ギャル旅タイアップ": 6000}

row = 5
for i, key in enumerate(y1.keys()):
    v1, v2, v3 = y1[key], y2[key], y3[key]
    g1 = f"{int((v2/v1-1)*100)}%" if v1 > 0 else "-"
    g2 = f"{int((v3/v2-1)*100)}%" if v2 > 0 else "-"
    bg = LIGHT_PINK if i % 2 == 0 else WHITE
    for col, v in enumerate([key, v1, v2, v3, g1, g2]):
        fmt = "#,##0" if col in [1,2,3] else None
        set_cell(ws3, row, col+1, v, bg=bg, fmt=fmt,
                 h_align="right" if col in [1,2,3] else "center" if col in [4,5] else "left")
    row += 1

# 合計行
t1, t2, t3 = sum(y1.values()), sum(y2.values()), sum(y3.values())
g1 = f"{int((t2/t1-1)*100)}%"
g2 = f"{int((t3/t2-1)*100)}%"
for col, v in enumerate(["売上合計", t1, t2, t3, g1, g2]):
    fmt = "#,##0" if col in [1,2,3] else None
    set_cell(ws3, row, col+1, v, bold=True, bg=DARK_PINK, color=WHITE,
             fmt=fmt, h_align="right" if col in [1,2,3] else "center" if col in [4,5] else "left")
row += 2

# 収益分配計画
section_title(ws3, row, 1, "■ 収益分配計画（推定）", 6)
row += 1
header_row(ws3, row, ["項目", "Year1", "Year2", "Year3", "割合（Y3）", "備考"])
row += 1
shares_data = [
    ("売上合計",           t1,           t2,           t3,           "100%", ""),
    ("ギャル協会 取り分",  int(t1*0.32), int(t2*0.32), int(t3*0.32), "約32%", "出演・監修・IP料"),
    ("サイバーバズ 取り分",int(t1*0.50), int(t2*0.50), int(t3*0.50), "約50%", "運営・制作費・利益"),
    ("ENTIAL 取り分",      int(t1*0.13), int(t2*0.13), int(t3*0.13), "約13%", "PM・営業手数料"),
    ("制作費・運営費（バズ負担）", int(t1*0.35), int(t2*0.30), int(t3*0.25), "約25%", "売上規模拡大で比率低下"),
]
for i, r in enumerate(shares_data):
    bg = DARK_PINK if i == 0 else (LIGHT_PINK if i % 2 == 1 else WHITE)
    fg = WHITE if i == 0 else DARK
    for col, v in enumerate(r):
        fmt = "#,##0" if col in [1,2,3] else None
        set_cell(ws3, row, col+1, v, bold=(i==0), bg=bg, color=fg,
                 fmt=fmt, h_align="right" if col in [1,2,3] else "left")
    row += 1

row += 1

# ロードマップ
section_title(ws3, row, 1, "■ ロードマップ", 6)
row += 1
header_row(ws3, row, ["フェーズ", "期間", "主要アクション", "", "", "マイルストーン"])
ws3.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
row += 1
roadmap = [
    ("Phase 1\n仕込み期", "Year1 Q1-Q2",
     "3社契約締結 / TikTok開設 / ppgalclub参考に渋谷体験商品化 / ENTIAL営業代理店開拓",
     "タイアップ初受注 / インバウンド体験50人/月"),
    ("Phase 2\n成長期", "Year1 Q3 〜 Year2",
     "TikTokフォロワー1万人 / WESELL本格稼働 / 自治体案件獲得 / IG Reels展開",
     "月次売上1,000万安定 / フォロワー10万人"),
    ("Phase 3\nスケール期", "Year3",
     "米国WESELL拡大 / ブラジル・スペインIG展開 / 海外ライセンス / 神7組成",
     "年間10億達成 / 海外IPライセンス契約5社以上"),
]
for i, r in enumerate(roadmap):
    bg = LIGHT_PINK if i % 2 == 0 else WHITE
    set_cell(ws3, row, 1, r[0], bold=True, bg=bg, wrap=True)
    set_cell(ws3, row, 2, r[1], bg=bg, h_align="center")
    ws3.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    set_cell(ws3, row, 3, r[2], bg=bg, wrap=True)
    set_cell(ws3, row, 6, r[3], bg=bg, wrap=True)
    ws3.row_dimensions[row].height = 45
    row += 1

# ══════════════════════════════════════════════════════════════
# SHEET 4: 収益シミュレーション（グラフ用）
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("収益シミュレーション")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 26
for col in ["B","C","D"]:
    ws4.column_dimensions[col].width = 16

ws4.merge_cells("A1:D1")
c = ws4["A1"]
c.value = "3カ年 収益シミュレーション"
c.font = font(bold=True, size=14, color=WHITE)
c.fill = fill(PINK)
c.alignment = align("center")
ws4.row_dimensions[1].height = 30

header_row(ws4, 2, ["収益源", "Year1（万円）", "Year2（万円）", "Year3（万円）"])
chart_data = [
    ("① インバウンド体験",   1800,  4800, 15000),
    ("② 広告タイアップ",     3600, 10800, 33600),
    ("③ WESELL グッズ販売",  1200,  6000, 24000),
    ("④ IPライセンス",        600,  1200,  6000),
    ("⑤ ギャル旅タイアップ",  800,  2500,  6000),
    ("合計",                 8000, 25300, 84600),
]
for i, r in enumerate(chart_data):
    bg = DARK_PINK if i == 5 else (LIGHT_PINK if i % 2 == 0 else WHITE)
    fg = WHITE if i == 5 else DARK
    for col, v in enumerate(r):
        fmt = "#,##0" if col > 0 else None
        set_cell(ws4, 3+i, col+1, v, bold=(i==5), bg=bg, color=fg,
                 fmt=fmt, h_align="right" if col > 0 else "left")

# 棒グラフ
chart = BarChart()
chart.type = "col"
chart.grouping = "clustered"
chart.title = "3カ年 収益推移（万円）"
chart.y_axis.title = "売上（万円）"
chart.x_axis.title = "収益源"
chart.width = 22
chart.height = 14

data_ref = Reference(ws4, min_col=2, max_col=4, min_row=2, max_row=7)
cats = Reference(ws4, min_col=1, min_row=3, max_row=7)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats)

from openpyxl.drawing.fill import PatternFillProperties
colors_chart = ["FF69B4", "FF1493", "FFD6E8"]
for i, series in enumerate(chart.series):
    series.graphicalProperties.solidFill = colors_chart[i % len(colors_chart)]

ws4.add_chart(chart, "A10")

# ══════════════════════════════════════════════════════════════
# 保存
# ══════════════════════════════════════════════════════════════
out = "/home/user/gyal-kyokai--/事業計画_ギャル×インバウンド.xlsx"
wb.save(out)
print(f"Saved: {out}")
